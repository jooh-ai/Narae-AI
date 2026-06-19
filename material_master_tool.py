# -*- coding: utf-8 -*-
"""
자재마스터 사양 분배 도구 (Python 버전)
================================================

M열(사양)의  [속성명:값][속성명:값] ...  데이터를 같은 행의
O~AH열 개별속성 칸에 분배합니다. 최종 결과는 각 칸에 '값만' 남고
(속성명 라벨은 제거), 어느 데이터도 들어가지 않은 개별속성 칸은
빈칸이 됩니다. 개별속성 열 위치는 그대로 유지됩니다.
어느 속성과도 매칭되지 않는 데이터는 ETC 칸에 모읍니다.
M열에 [속성명:값] 항목이 없는 행(빈 셀 또는 '_' 같은 값)은
MAKER P/N 외에는 데이터가 없으므로, MAKER P/N 칸만 '값'으로 남기고
그 행의 나머지 개별속성 칸은 (어떤 라벨이든) 모두 빈칸으로 정리합니다.
MAKER P/N 칸은 'MAKER P/N:12345' 라벨 형태이거나 L열 값과 동일한 칸으로
식별하며, 식별하지 못하면 데이터 손실을 막기 위해 그 행은 건드리지
않고 경고만 남깁니다.

주의: 이 도구는 개별속성 칸의 라벨을 값으로 바꾸거나 비우므로,
반드시 '라벨이 들어있는 원본 마스터'에 대해 실행해야 합니다.
(이미 처리된 결과 파일에 다시 실행하면 매칭이 되지 않습니다.)

셀 서식(폰트·색상·테두리·열 너비 등)은 openpyxl이 그대로 보존하며,
새로 기입한 셀은 빨간 글자색으로 강조할 수 있습니다.

사용법
------
1) 한 번만:  터미널에서  ->  pip install openpyxl
2) 아래 ##### 설정 ##### 블록의 값을 본인 파일에 맞게 수정
3) VSCode에서 F5(또는 '실행')로 이 파일을 실행

원본 파일은 수정하지 않고, 결과는 새 파일로 저장됩니다.
"""

import os
import sys
import re
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import column_index_from_string
except ImportError:
    print("✗ openpyxl 라이브러리가 필요합니다. 터미널에서 다음을 실행하세요:")
    print("    pip install openpyxl")
    sys.exit(1)


# ═══════════════════════════════════════════════
#                    설정
# ═══════════════════════════════════════════════
# 입력 파일 경로 (이 .py 파일 기준 상대경로 또는 절대경로)
INPUT_FILE      = "자재마스터.xlsx"

# 수정할 시트명. 빈 문자열("")이면 '위례'가 포함된 시트를 자동 선택, 없으면 첫 시트
SHEET_NAME      = ""

# 헤더 행 번호 (이 행 다음부터 데이터로 처리)
HEADER_ROW      = 1

# 컬럼 지정 (엑셀 알파벳)
SPEC_COL        = "M"    # 사양 컬럼
PN_COL          = "L"    # MAKER P/N 컬럼 (사양이 비어있을 때 보완용)
ATTR_START_COL  = "O"    # 개별속성 시작열
ATTR_END_COL    = "AH"   # 개별속성 끝열

# 새로 기입한 셀을 빨간 글자색으로 강조할지 여부
HIGHLIGHT_RED   = True

# M열 사양이 없는 행을 L열(MAKER P/N) 값으로 보완할지 여부.
# False(기본)면 그런 행은 전혀 건드리지 않고 그대로 둔다(이미 채워둔 값 보존).
PN_FALLBACK     = False

# 결과 파일명. 빈 문자열("")이면 "입력파일명_YYMMDD_HHMM.xlsx" 로 자동 생성
OUTPUT_FILE     = ""
# ═══════════════════════════════════════════════


RED_ARGB = "FFFF0000"
MAKER_PN_NORM = "MAKERPN"


# ─────────────────────────────────────────────
# 헬퍼
# ─────────────────────────────────────────────
def cell_text(cell):
    """셀 값을 문자열로 반환 (None → '')."""
    v = cell.value
    if v is None:
        return ""
    return str(v)


def is_empty(v):
    """공백/제로폭 문자만 있으면 빈 값으로 간주."""
    if v is None:
        return True
    s = str(v)
    cleaned = re.sub(r"[\s ​‌‍﻿]+", "", s)
    return cleaned == ""


def norm_label(s):
    """속성명 정규화: ':' 앞부분을 대문자화 + 공백·점·슬래시 제거.
    'MAKER P/N' → 'MAKERPN', 'CONN. SIZE' → 'CONNSIZE'."""
    t = "" if s is None else str(s)
    m = re.search(r"[:：]", t)
    if m:
        t = t[: m.start()]
    return re.sub(r"[\s.\/]", "", t.upper())


def norm_value(s):
    """매칭/중복 비교용 값 정규화 (공백·언더스코어 제거, 소문자)."""
    return re.sub(r"[\s_]", "", ("" if s is None else str(s)).lower())


def label_part(text):
    """셀/세그먼트 텍스트에서 ':' 이전의 라벨 부분만 추출(trim)."""
    t = str(text)
    m = re.search(r"[:：]", t)
    if m:
        t = t[: m.start()]
    return t.strip()


def apply_red_font(cell):
    """기존 폰트 속성을 보존하면서 글자색만 빨강으로 변경."""
    f = cell.font
    cell.font = Font(
        name=f.name, size=f.size, bold=f.bold, italic=f.italic,
        vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
        color=RED_ARGB,
    )


def write_cell(cell, value, highlight):
    cell.value = value
    if highlight:
        apply_red_font(cell)


# ─────────────────────────────────────────────
# 사양 파싱
# ─────────────────────────────────────────────
def push_segment(content, entries, unlabeled):
    """대괄호 내용 하나를 라벨/값으로 분해."""
    t = (content or "").strip()
    if not t:
        return
    m = re.search(r"[:：]", t)
    if m:
        label = t[: m.start()].strip()
        value = t[m.start() + 1:].strip()
        if label:
            entries.append({"label": label, "value": value, "raw": t})
            return
    unlabeled.append(t)


def parse_spec(spec):
    """M열 사양 문자열을 파싱.
    반환: (entries, unlabeled)
      entries  : [{label, value, raw}]  - [라벨:값] 형태
      unlabeled: [str]                  - 라벨 없는 미매칭 후보
    대괄호 밖의 '_Vendor - EMERSON' 같은 echo(중복)는 자동으로 버린다."""
    entries = []
    unlabeled = []
    text = "" if spec is None else str(spec)

    # 1) 대괄호 내용 추출
    for m in re.finditer(r"\[([^\[\]]*)\]", text):
        push_segment(m.group(1), entries, unlabeled)

    # 2) 대괄호 밖 잔여 텍스트
    remainder = re.sub(r"\[[^\[\]]*\]", " ", text).strip()
    remainder = re.sub(r"^[\s_]+|[\s_]+$", "", remainder)
    if remainder:
        known = set()
        for e in entries:
            if e["value"]:
                known.add(norm_value(e["value"]))
            known.add(norm_value(e["raw"]))
        for u in unlabeled:
            known.add(norm_value(u))
        if norm_value(remainder) not in known:
            push_segment(remainder, entries, unlabeled)

    return entries, unlabeled


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def main():
    # 이 .py 파일이 있는 폴더. Jupyter/대화형 창에서는 __file__ 이 없으므로
    # 현재 작업 디렉터리(cwd)로 대체한다.
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        base_dir = os.getcwd()
    in_path = INPUT_FILE if os.path.isabs(INPUT_FILE) else os.path.join(base_dir, INPUT_FILE)

    if not os.path.exists(in_path):
        print(f"✗ 입력 파일을 찾을 수 없습니다: {in_path}")
        print("  설정 블록의 INPUT_FILE 값을 확인하세요.")
        sys.exit(1)

    print(f"▶ 파일 로드 중: {in_path}")
    keep_vba = in_path.lower().endswith(".xlsm")
    wb = load_workbook(in_path, keep_vba=keep_vba)

    # 시트 선택
    if SHEET_NAME:
        if SHEET_NAME not in wb.sheetnames:
            print(f"✗ 시트 '{SHEET_NAME}'를 찾을 수 없습니다. 사용 가능한 시트: {wb.sheetnames}")
            sys.exit(1)
        ws = wb[SHEET_NAME]
    else:
        chosen = next((n for n in wb.sheetnames if "위례" in n), wb.sheetnames[0])
        ws = wb[chosen]
    print(f"  · 대상 시트: '{ws.title}'")

    spec_col = column_index_from_string(SPEC_COL)
    pn_col = column_index_from_string(PN_COL)
    attr_start = column_index_from_string(ATTR_START_COL)
    attr_end = column_index_from_string(ATTR_END_COL)
    if attr_end < attr_start:
        print("✗ 개별속성 시작열/끝열 설정을 확인하세요.")
        sys.exit(1)

    max_row = ws.max_row
    total_rows = max_row - HEADER_ROW
    print(f"  · 데이터 {total_rows}행, 개별속성 {ATTR_START_COL}~{ATTR_END_COL}열")

    rows_with_spec = 0       # 사양(=[라벨:값] 항목)이 있던 행
    value_cells = 0          # 데이터(값)를 기입한 칸
    cleared_cells = 0        # 라벨/빈 데이터를 비운 칸
    etc_rows = 0             # ETC 칸에 미매칭 데이터를 기입한 행
    pn_fallback_rows = 0     # 사양 없는 행에서 L열 값으로 MAKER P/N을 보완한 칸
    no_etc_warn = 0          # 미매칭인데 ETC 칸을 못 찾은 행
    specless_cleaned = 0     # 사양 없는 행: MAKER P/N만 남기고 정리한 행
    specless_skipped = 0     # 사양 없는 행: MAKER P/N을 식별 못해 건드리지 않은 행
    sample_unmatched = []
    sample_skipped = []

    for r in range(HEADER_ROW + 1, max_row + 1):
        spec_val = cell_text(ws.cell(row=r, column=spec_col))
        entries, unlabeled = parse_spec(spec_val)

        # ── (A) 사양에 [라벨:값] 항목이 없는 행(빈 셀 또는 '_' 같은 무의미 값) ──
        #     이런 행은 MAKER P/N 외에는 데이터가 없으므로, MAKER P/N 칸만 '값'으로
        #     남기고 그 행의 나머지 칸은 (어떤 라벨이든) 모두 비운다.
        #     MAKER P/N 칸 식별: ① 'MAKER P/N:값' 라벨 형태  ② L열 값과 동일한 칸
        if not entries:
            pn_l_raw = cell_text(ws.cell(row=r, column=pn_col)).strip()
            pn_l = norm_value(pn_l_raw)
            nonempty = []           # [(col, text)]
            keep_col = None         # MAKER P/N 데이터가 들어있는 칸
            keep_val = None         # 그 칸에 남길 값 (None이면 비움)
            keep_is_fallback = False
            for c in range(attr_start, attr_end + 1):
                v = cell_text(ws.cell(row=r, column=c))
                if is_empty(v):
                    continue
                s = str(v)
                nonempty.append((c, s))
                if keep_col is not None:
                    continue
                ci = next((i for i, ch in enumerate(s) if ch in ":："), -1)
                val = s[ci + 1:].strip() if ci >= 0 else ""  # ':' 뒤의 값
                if norm_label(s) == MAKER_PN_NORM:
                    # 'MAKER P/N:12345' → '12345'.  라벨만 있으면 L열 값(보완) 또는 비움
                    keep_col = c
                    if val:
                        keep_val = val
                    elif PN_FALLBACK and pn_l_raw:
                        keep_val = pn_l_raw
                        keep_is_fallback = True
                    else:
                        keep_val = None
                elif pn_l and norm_value(s) == pn_l:
                    # 라벨 없는 MAKER P/N 값(L열 값과 동일한 칸)
                    keep_col = c
                    keep_val = val if val else s.strip()

            if keep_col is None:
                # MAKER P/N을 식별하지 못함 → 데이터 손실 방지를 위해 건드리지 않음
                if nonempty:
                    specless_skipped += 1
                    if len(sample_skipped) < 10:
                        sample_skipped.append(f"{r}행")
                continue

            # MAKER P/N 칸은 값만 남기고(없으면 비움), 나머지 칸은 모두 비운다
            for c, s in nonempty:
                if c == keep_col:
                    if keep_val is None:
                        ws.cell(row=r, column=c).value = None
                        cleared_cells += 1
                    elif s.strip() != keep_val:
                        write_cell(ws.cell(row=r, column=c), keep_val, HIGHLIGHT_RED)
                        value_cells += 1
                        if keep_is_fallback:
                            pn_fallback_rows += 1
                else:
                    ws.cell(row=r, column=c).value = None
                    cleared_cells += 1
            specless_cleaned += 1
            continue

        # ── 사양이 있는 행: 개별속성 라벨 → 열 맵 + ETC 열 식별 ──
        #     (라벨은 값으로 덮어쓰거나 지우기 전에 먼저 읽어 둔다)
        label_to_col = {}    # norm → col
        etc_col = 0
        attr_cols = []       # 이 행에서 라벨이 있던 개별속성 열들
        for c in range(attr_start, attr_end + 1):
            txt = cell_text(ws.cell(row=r, column=c))
            if is_empty(txt):
                continue
            norm = norm_label(txt)
            if not norm:
                continue
            attr_cols.append(c)
            if norm not in label_to_col:
                label_to_col[norm] = c
            if "ETC" in norm:
                etc_col = c

        # (B) 사양 분배: 각 개별속성 열에 채울 '값'을 결정 (col → value)
        rows_with_spec += 1
        col_value = {}
        unmatched = []
        for e in entries:
            norm = norm_label(e["label"])
            if "ETC" in norm:
                if e["value"]:
                    unmatched.append(e["value"])
                continue
            target = label_to_col.get(norm)
            if target is not None and target != etc_col:
                if e["value"]:
                    col_value[target] = e["value"]
            else:
                unmatched.append(e["raw"])
        for u in unlabeled:
            unmatched.append(u)
        # 미매칭 데이터는 ETC 칸으로
        if unmatched:
            if etc_col:
                col_value[etc_col] = ", ".join(unmatched)
                etc_rows += 1
            else:
                no_etc_warn += 1
                if len(sample_unmatched) < 10:
                    sample_unmatched.append(f"{SPEC_COL}{r}: {' | '.join(unmatched)}")

        # (C) 개별속성 칸 정리 (사양이 있는 행, 열 위치는 그대로 유지):
        #     - 결정된 값이 있으면 '값만' 기입 (속성명 라벨 제거)
        #     - 값이 없으면(라벨만 있던 칸) 빈칸으로
        for c in attr_cols:
            if c in col_value:
                write_cell(ws.cell(row=r, column=c), col_value[c], HIGHLIGHT_RED)
                value_cells += 1
            else:
                ws.cell(row=r, column=c).value = None
                cleared_cells += 1

        if (r - HEADER_ROW) % 500 == 0:
            print(f"  ... {r - HEADER_ROW}/{total_rows} 행 처리")

    print(f"  → 처리 완료: 사양 보유 {rows_with_spec}행, 사양 없는 행 정리 {specless_cleaned}행, "
          f"값 기입 {value_cells}칸, 빈칸 정리 {cleared_cells}칸, ETC 모음 {etc_rows}행, L열 보완 {pn_fallback_rows}칸")
    if specless_skipped:
        print(f"  ⚠ 사양 없는 행 중 MAKER P/N 값(=L열 값 또는 'MAKER P/N:' 라벨)을 못 찾아 "
              f"건드리지 않은 행 {specless_skipped}건")
        for s in sample_skipped:
            print(f"      · {s}")
    if no_etc_warn:
        print(f"  ⚠ ETC 속성 칸을 못 찾아 미매칭 데이터를 기입 못한 행 {no_etc_warn}건")
        for s in sample_unmatched:
            print(f"      · {s}")

    # 저장
    if OUTPUT_FILE:
        out_path = OUTPUT_FILE if os.path.isabs(OUTPUT_FILE) else os.path.join(base_dir, OUTPUT_FILE)
    else:
        base = os.path.splitext(os.path.basename(in_path))[0]
        stamp = datetime.now().strftime("%y%m%d_%H%M")
        ext = ".xlsm" if keep_vba else ".xlsx"
        out_path = os.path.join(base_dir, f"{base}_{stamp}{ext}")

    wb.save(out_path)
    print(f"✅ 저장 완료: {out_path}")


if __name__ == "__main__":
    main()
