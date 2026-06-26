"""날씨 로더 — 엑셀3-1(외부망 크롤링) 파싱 → 입찰 적용 대기압 산정.

엑셀3-1 레이아웃(성남비행장, 7일 × 3시간대):
  A1            : 캡처 시각
  'Pressure'    : 섹션 시작 → 다음 헤더행(시간대 0,3,…,21 + '중위'+'최소'), 이후 일자별 행
  'Tempereture' : 온도 섹션(동일 구조, '중위'+'취약')

적용 대기압 = (해당 일 또는 전체 '중위' 대기압) + 위치보정(−8mbar; 성남비행장↔발전소).
"""
from __future__ import annotations

from dataclasses import dataclass, field

from . import constants as C


@dataclass
class WeatherForecast:
    capture: str | None
    days: list[str]
    times: list[int]
    pressure_median: dict[str, float]
    pressure_min: dict[str, float] = field(default_factory=dict)
    pressure_grid: dict[str, list[float]] = field(default_factory=dict)   # day → 시간대별 대기압
    temp_median: dict[str, float] = field(default_factory=dict)


def _grid(path: str):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [[c.value for c in r] for r in ws.iter_rows()]


def _find_label(rows, label: str) -> int | None:
    for i, r in enumerate(rows):
        for v in r:
            if isinstance(v, str) and v.strip() == label:
                return i
    return None


def _parse_block(rows, start: int):
    """섹션 시작행 이후의 (중위dict, 최소dict, grid, days, times) 파싱."""
    hdr = None
    for i in range(start + 1, len(rows)):
        if any(isinstance(v, str) and v.strip() == "중위" for v in rows[i]):
            hdr = i
            break
    if hdr is None:
        return {}, {}, {}, [], []          # 5-tuple (grid 포함) — 헤더 없을 때 안전 반환
    header = rows[hdr]
    med_c = next(j for j, v in enumerate(header) if isinstance(v, str) and v.strip() == "중위")
    min_c = next((j for j, v in enumerate(header)
                  if isinstance(v, str) and v.strip() in ("최소", "취약")), None)
    time_cols = [j for j, v in enumerate(header)
                 if isinstance(v, (int, float)) and not isinstance(v, bool)]
    times = [int(header[j]) for j in time_cols]
    days: list[str] = []
    med: dict[str, float] = {}
    mn: dict[str, float] = {}
    grid: dict[str, list[float]] = {}
    for i in range(hdr + 1, len(rows)):
        label = rows[i][0] if rows[i] else None
        if not isinstance(label, str) or not label.strip():
            break
        if label.strip() in ("Pressure", "Tempereture", "Temperature"):
            break
        days.append(label)
        row = rows[i]
        if med_c < len(row) and isinstance(row[med_c], (int, float)):
            med[label] = float(row[med_c])
        if min_c is not None and min_c < len(row) and isinstance(row[min_c], (int, float)):
            mn[label] = float(row[min_c])
        grid[label] = [float(row[j]) if j < len(row) and isinstance(row[j], (int, float))
                       else None for j in time_cols]
    return med, mn, grid, days, times


def load_excel3_1(path: str) -> WeatherForecast:
    """엑셀3-1 파일 파싱."""
    rows = _grid(path)
    capture = rows[0][0] if rows and rows[0] else None
    p_i = _find_label(rows, "Pressure")
    t_i = _find_label(rows, "Tempereture") or _find_label(rows, "Temperature")
    if p_i is not None:
        p_med, p_min, p_grid, days, times = _parse_block(rows, p_i)
    else:
        p_med, p_min, p_grid, days, times = {}, {}, {}, [], []
    t_med = _parse_block(rows, t_i)[0] if t_i is not None else {}
    return WeatherForecast(capture=str(capture) if capture is not None else None,
                           days=days, times=times, pressure_median=p_med,
                           pressure_min=p_min, pressure_grid=p_grid, temp_median=t_med)


def applied_pressure(fc: WeatherForecast, *, day: str | None = None,
                     offset: float = C.WEATHER_SITE_OFFSET, aggregate: str = "mean") -> float:
    """입찰 적용 대기압(mbar) = 중위 대기압 + 위치보정(−8).

    day 지정 시 그 날의 중위, 아니면 전체 중위의 평균(mean) 또는 중앙값(median)."""
    import statistics
    if day is not None:
        if day not in fc.pressure_median:
            raise ValueError(
                f"'{day}' 날짜의 중위 대기압이 없습니다 (가용: {fc.days})")
        base = fc.pressure_median[day]
    else:
        vals = [v for v in fc.pressure_median.values() if isinstance(v, (int, float))]
        if not vals:
            raise ValueError("중위 대기압 데이터가 없습니다")
        base = statistics.mean(vals) if aggregate == "mean" else statistics.median(vals)
    return base + offset
