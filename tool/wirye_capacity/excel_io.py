"""엑셀 파일 열기 공통 처리 — 사내 보안(DRM)·손상 파일을 사람이 읽을 수 있는 오류로.

openpyxl 은 암호화·손상 파일에서 `BadZipFile: File is not a zip file` 를 던지는데,
현장에서는 "사내 문서 보안이 걸린 파일"이 가장 흔한 원인이다. 원인·해결책을 담은
ExcelOpenError 로 바꿔 GUI 팝업/CLI 에 그대로 노출한다.

오류 메시지에는 **전체 경로·파일 크기**를 함께 싣는다. 파일명만 보여주면
"사용자가 지정한 파일"인지 "프로그램에 번들된 템플릿"인지 구분할 수 없어
원인 파악이 불가능하다. 번들 템플릿이면 조치 방법이 완전히 다르다(재빌드).
"""
from __future__ import annotations

import sys
from pathlib import Path


class ExcelOpenError(RuntimeError):
    """엑셀 파일을 열 수 없음(보안/손상). 메시지에 해결 방법 포함."""


def _is_bundled(p: Path) -> bool:
    """PyInstaller 로 exe 안에 번들된 파일인지 (_MEIPASS 하위인지)."""
    meipass = getattr(sys, "_MEIPASS", None)
    if not meipass:
        return False
    try:
        p.resolve().relative_to(Path(meipass).resolve())
        return True
    except (ValueError, OSError):
        return False


def _detail(p: Path) -> str:
    """진단에 필요한 사실만 — 전체 경로와 크기."""
    try:
        size = f"{p.stat().st_size:,} 바이트"
    except OSError:
        size = "크기 확인 불가"
    return f"경로: {p}\n크기: {size}"


def _remedy(p: Path) -> str:
    """번들 템플릿과 사용자 지정 파일은 조치가 다르다."""
    if _is_bundled(p):
        return ("이 파일은 프로그램에 포함된 입찰 양식 템플릿입니다.\n"
                "사용자 파일이 아니므로 보안 해제 대상이 아닙니다.\n\n"
                "해결 순서:\n"
                "  1) 최신 코드를 받고 다시 빌드하세요\n"
                "       git pull  →  pyinstaller --noconfirm wirye_tool.spec\n"
                "  2) 그래도 같으면 소스의 템플릿이 손상된 경우입니다\n"
                "       git checkout -- wirye_capacity/templates/\n"
                "  3) 급하면 정상 엑셀3 양식 파일을 직접 지정할 수 있습니다\n"
                "       설정파일(~/.wirye_tool.json)의 template_path")
    return ("원인: 사내 문서 보안(DRM)이 걸려 있거나, 다운로드 중 파일이 손상되었습니다.\n\n"
            "해결: 파일을 Excel 에서 열어 [파일 → 다른 이름으로 저장]으로 "
            "보안이 해제된 사본(.xlsx)을 만든 뒤 그 파일을 지정하세요.")


def load_workbook_safe(path, **kwargs):
    """openpyxl.load_workbook 래퍼 — 보안/손상 파일을 ExcelOpenError 로 변환."""
    import zipfile

    from openpyxl import load_workbook
    p = Path(path)
    if not p.exists():
        raise ExcelOpenError(f"파일이 없습니다:\n{p}")
    if p.stat().st_size == 0:
        raise ExcelOpenError(
            f"엑셀 파일이 비어 있습니다(0바이트): {p.name}\n\n{_detail(p)}\n\n{_remedy(p)}")
    try:
        return load_workbook(p, **kwargs)
    except zipfile.BadZipFile as e:
        raise ExcelOpenError(
            f"엑셀 파일을 열 수 없습니다: {p.name}\n\n{_detail(p)}\n\n{_remedy(p)}") from e
    except Exception as e:   # noqa: BLE001 — 암호 보호 등
        raise ExcelOpenError(
            f"엑셀 파일을 열 수 없습니다: {p.name}\n\n{_detail(p)}\n\n{e}") from e
