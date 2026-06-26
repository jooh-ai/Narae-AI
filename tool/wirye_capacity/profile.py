"""온도 Profile 생성 — 최종 산출물 (엑셀3/엑셀4 '온도 Profile' 형식).

Mode3 기준으로 −20~40°C 각 온도의 이론/현실화 GT·ST·CC를 산출하고 .xlsx로 출력한다.

  CC 이론(Gross)  = base_cc(T) × (1.028/Deg) / P_corr(대기압) + W(IGV)
  GT 이론(Gross)  = base_gt(T) × (1.028/Deg) / P_corr(대기압) + W·GT비
  ST 이론         = CC 이론 − GT 이론
  CC 현실화(Gross)= min( CC 이론 + 보정값(온도구간), 상한 Gross )
  CC 현실화(Net)  = min( CC 현실화 Gross − 소내전력(10), 462 )
  GT 현실화 = GT 이론,  ST 현실화 = CC 현실화 − GT 이론  (보정값은 CC에만)
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from . import constants as C
from .correction import applied_correction, realized_net
from .theory import TheoryEngine, igv_turnup

# 기본 엑셀3 템플릿 (최종 입찰 양식)
DEFAULT_TEMPLATE = Path(__file__).parent / "templates" / "excel3_profile_template.xlsx"


@dataclass
class ProfileRow:
    temp: int          # CIT (°C)
    w: float           # IGV turn-up (MW)
    gt_theory: float
    st_theory: float
    cc_theory: float       # Gross, IGV 반영
    correction: float      # 온도구간 적용 보정값
    cc_real_gross: float
    cc_real_net: float     # 입찰값 (≤462)
    gt_real: float
    st_real: float


def build_profile(engine: TheoryEngine, correction_table: dict,
                  pressure: float = C.REF_PRESSURE, deg: float = C.DEFAULT_DEG,
                  temps: list[int] | None = None, corrector=None) -> list[ProfileRow]:
    """온도별 이론/현실화 Profile(Mode3) 생성.

    correction_table: store.correction_table() / aggregate_bins() 결과 (구간 방식).
    corrector: 있으면 보정값을 corrector(CIT)로 산출(연속곡선 등). 없으면 구간 평균.
    """
    if temps is None:
        temps = list(range(-20, 41))
    k_factor = lambda: (C.REF_DEG / deg) / engine.p_corr(pressure)
    k = k_factor()
    rows: list[ProfileRow] = []
    for t in temps:
        w = igv_turnup(t)
        cc_theory = engine.base_cc(t) * k + w
        gt_theory = engine.base_gt(t) * k + w * C.GT_RATIO
        st_theory = cc_theory - gt_theory
        corr = corrector(t) if corrector is not None else applied_correction(t, correction_table)
        cc_real_gross = min(cc_theory + corr, C.BID_CAP_GROSS)
        cc_real_net = realized_net(cc_theory, corr)         # min(gross-aux, 462)
        gt_real = gt_theory                                  # 보정값은 CC에만
        st_real = cc_real_gross - gt_real
        rows.append(ProfileRow(
            temp=t, w=w, gt_theory=gt_theory, st_theory=st_theory, cc_theory=cc_theory,
            correction=corr, cc_real_gross=cc_real_gross, cc_real_net=cc_real_net,
            gt_real=gt_real, st_real=st_real))
    return rows


# 출력 컬럼 (헤더, ProfileRow 속성)
_COLUMNS = [
    ("외기온도(CIT,°C)", "temp"),
    ("GT 이론", "gt_theory"),
    ("ST 이론", "st_theory"),
    ("CC 이론(Gross)", "cc_theory"),
    ("보정값", "correction"),
    ("GT 현실화", "gt_real"),
    ("ST 현실화", "st_real"),
    ("★CC 현실화(Gross)", "cc_real_gross"),
    ("★CC 현실화(Net,입찰)", "cc_real_net"),
]


def write_xlsx(profile: list[ProfileRow], path: str, *,
               pressure: float = C.REF_PRESSURE, deg: float = C.DEFAULT_DEG) -> str:
    """Profile을 엑셀3 형식 '온도 Profile' 시트로 저장."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "온도 Profile"

    title = (f"위례열병합발전소  Mode3_AOH1000  온도별 공급가능용량  "
             f"(대기압 {pressure:g} mbar · Deg {deg:g} · 입찰 Net ≤ {C.BID_CAP_NET:g} MW)")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUMNS))

    ws["A2"] = "입력"
    ws["B2"] = "대기압(mbar)"
    ws["C2"] = pressure
    ws["D2"] = "Deg"
    ws["E2"] = deg
    for c in ("A2", "B2", "D2"):
        ws[c].font = Font(bold=True)
    ws["C2"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["E2"].fill = PatternFill("solid", fgColor="E2EFDA")

    hdr = 4
    for j, (label, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=hdr, column=j, value=label)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    for i, row in enumerate(profile, start=hdr + 1):
        for j, (_, attr) in enumerate(_COLUMNS, start=1):
            val = getattr(row, attr)
            cell = ws.cell(row=i, column=j, value=round(val, 2) if isinstance(val, float) else val)
            if j >= 2:
                cell.number_format = "0.00"

    widths = [16, 10, 10, 14, 9, 11, 11, 17, 19]
    for j, wdt in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=hdr, column=j).column_letter].width = wdt
    ws.freeze_panes = ws.cell(row=hdr + 1, column=2)

    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 엑셀3 템플릿 채우기 (최종 입찰 양식 — 기존 엑셀3 그대로)
#
# 기존 워크플로(엑셀2 P,Q → 엑셀3 Mode3 B,C 붙여넣기)를 자동화한다. Tool이 계산한
# 현실화 Mode3 GT/ST(Gross)를 Mode3!B5:C65 에 기입하면, 엑셀3의 기존 수식이 6개 모드·
# 온도 Profile·NET·열전비 파생을 자동 산출한다(Excel에서 열면 재계산).
# ─────────────────────────────────────────────────────────────────────────────

# 엑셀3 'Mode3' 시트: A=온도, B=GT Gross, C=ST Gross (행5 = −20°C … 행65 = 40°C)
_MODE3_FIRST_ROW = 5
# 엑셀3 '온도 Profile' 날씨(Pressure) 블록: P=일자, Q~X=시간대, Y=중위, Z=최소 (행6~12)
_WX = {"sheet": "온도 Profile", "first_row": 6, "day_col": 16,
       "first_time_col": 17, "median_col": 25, "min_col": 26}


def _fill_weather(wb, forecast) -> None:
    """온도 Profile 날씨(Pressure) 블록에 엑셀3-1 예보를 기입 (M2 적용대기압 자동 갱신)."""
    if _WX["sheet"] not in wb.sheetnames:
        return
    ws = wb[_WX["sheet"]]
    for i, day in enumerate(forecast.days[:7]):
        r = _WX["first_row"] + i
        ws.cell(row=r, column=_WX["day_col"], value=day)
        for j, val in enumerate(forecast.pressure_grid.get(day, [])):
            if val is not None:
                ws.cell(row=r, column=_WX["first_time_col"] + j, value=val)
        if day in forecast.pressure_median:
            ws.cell(row=r, column=_WX["median_col"], value=forecast.pressure_median[day])
        if day in forecast.pressure_min:
            ws.cell(row=r, column=_WX["min_col"], value=forecast.pressure_min[day])


def fill_excel3_template(output_path: str, *, engine: TheoryEngine, correction_table: dict,
                         pressure: float = C.REF_PRESSURE, deg: float = C.DEFAULT_DEG,
                         forecast=None, template_path: str | Path = DEFAULT_TEMPLATE,
                         mode3_sheet: str = "Mode3", corrector=None) -> str:
    """현실화 Mode3 GT/ST를 엑셀3 템플릿 Mode3!B5:C65 에 채워 최종 입찰 파일 생성.

    pressure: 입찰 적용 대기압(보통 weather.applied_pressure = 중위−8). 현실화값에 반영됨.
    forecast: 있으면 온도 Profile 날씨 블록도 채움(M2 적용대기압 갱신).
    corrector: 있으면 연속곡선 등으로 보정값 산출(없으면 구간 평균).
    반환: output_path. (Excel에서 열면 6모드·온도Profile이 수식으로 재계산됨)
    """
    from openpyxl import load_workbook

    rows = build_profile(engine, correction_table, pressure=pressure, deg=deg,
                         temps=list(range(-20, 41)), corrector=corrector)
    wb = load_workbook(template_path)            # 수식·서식 보존
    m3 = wb[mode3_sheet]
    for i, row in enumerate(rows):
        rr = _MODE3_FIRST_ROW + i
        m3.cell(row=rr, column=2, value=row.gt_real)   # B = GT Gross (현실화)
        m3.cell(row=rr, column=3, value=row.st_real)   # C = ST Gross (= CC현실화 − GT)
    if forecast is not None:
        _fill_weather(wb, forecast)
    # Excel에서 열 때 6모드·온도Profile 수식이 재계산되도록
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    wb.save(output_path)
    return output_path
