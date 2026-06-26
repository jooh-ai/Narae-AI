"""검증(시운전) 모드 — 기준(엑셀) ↔ Tool 단계별 대조 리포트.

기존 검증된 엑셀 산출물을 "정답"으로, Tool 출력을 온도별로 대조해 허용오차(±0.5 MW) 합격 여부를
판정한다. 일회성 시운전이자 상시 회귀 도구.

reference 형식: {온도: {field: 값}}  (field 예: cc_theory, cc_real_gross, cc_real_net)
  - ref_from_profile(): Tool ProfileRow 목록 → reference dict
  - read_reference_xlsx(): 저장된 Profile .xlsx 읽기 (Tool 형식 또는 엑셀4 Profile 형식)
"""
from __future__ import annotations

from dataclasses import dataclass, field

from . import constants as C

DEFAULT_FIELDS = ["cc_theory", "cc_real_gross", "cc_real_net"]


@dataclass
class VerifyReport:
    tol: float
    rows: list[dict] = field(default_factory=list)   # {temp, field, tool, ref, diff, ok}

    @property
    def failures(self) -> list[dict]:
        return [r for r in self.rows if not r["ok"]]

    @property
    def passed(self) -> bool:
        return len(self.rows) > 0 and not self.failures

    @property
    def max_abs_diff(self) -> float:
        return max((abs(r["diff"]) for r in self.rows), default=0.0)

    def summary(self) -> str:
        ok = len(self.rows) - len(self.failures)
        verdict = "PASS ✅" if self.passed else "FAIL ❌"
        return (f"{verdict} — {ok}/{len(self.rows)} 합격, "
                f"최대 차이 {self.max_abs_diff:.3f} MW (허용 ±{self.tol} MW)")


def ref_from_profile(rows, fields: list[str] = DEFAULT_FIELDS) -> dict[int, dict]:
    """Tool ProfileRow 목록 → reference dict."""
    return {r.temp: {f: getattr(r, f) for f in fields} for r in rows}


def compare_profile(tool_rows, reference: dict[int, dict], tol: float = 0.5,
                    fields: list[str] = DEFAULT_FIELDS) -> VerifyReport:
    """Tool Profile ↔ reference 대조. 온도×field 별 차이/합격 판정."""
    rep = VerifyReport(tol=tol)
    for r in tool_rows:
        ref = reference.get(r.temp)
        if not ref:
            continue
        for f in fields:
            rv = ref.get(f)
            if rv is None or not hasattr(r, f):
                continue
            tv = getattr(r, f)
            diff = tv - rv
            rep.rows.append({"temp": r.temp, "field": f, "tool": tv, "ref": rv,
                             "diff": diff, "ok": abs(diff) <= tol})
    return rep


# 저장된 Profile 읽기용 컬럼 매핑 (열 번호는 1-base)
_TOOL_LAYOUT = {"first_row": 5, "temp": 1, "cc_theory": 4, "cc_real_gross": 8, "cc_real_net": 9}
# 엑셀4 'Mode3' Profile: A=온도(row6~66), D=CC이론, G=CC현실화(Gross)
_EXCEL4_LAYOUT = {"first_row": 6, "last_row": 66, "temp": 1, "cc_theory": 4, "cc_real_gross": 7}

LAYOUTS = {"tool": _TOOL_LAYOUT, "excel4": _EXCEL4_LAYOUT}


def read_reference_xlsx(path: str, layout: str = "tool", sheet: str | None = None) -> dict[int, dict]:
    """Profile .xlsx → reference dict.

    layout='tool'  : Tool write_xlsx 형식
    layout='excel4': 엑셀4 'Mode3' Profile 형식 (A온도/D CC이론/G CC현실화)
    """
    from openpyxl import load_workbook
    lay = LAYOUTS[layout]
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    fields = [k for k in ("cc_theory", "cc_real_gross", "cc_real_net") if k in lay]
    ref: dict[int, dict] = {}
    last = lay.get("last_row", ws.max_row)
    for row in range(lay["first_row"], last + 1):
        t = ws.cell(row=row, column=lay["temp"]).value
        if not isinstance(t, (int, float)) or isinstance(t, bool):
            continue
        entry = {}
        for f in fields:
            v = ws.cell(row=row, column=lay[f]).value
            if isinstance(v, (int, float)):
                entry[f] = float(v)
        ref[int(round(t))] = entry
    return ref
