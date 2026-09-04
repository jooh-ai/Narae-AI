"""날씨 로더(Phase 5) 검증 — 엑셀3-1 형식 합성 픽스처로 파싱·적용대기압 확인."""
import pytest

from wirye_capacity import constants as C
from wirye_capacity.weather import applied_pressure, load_excel3_1


def _make_excel3_1(path):
    """엑셀3-1 레이아웃(Pressure/Tempereture 섹션) 합성 픽스처."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "2026-04-15 18:09:12"
    ws["A2"] = "Update time :"
    ws["B2"] = "17:00"
    # Pressure 섹션
    ws["A5"] = "Pressure"
    times = [0, 3, 6, 9, 12, 15, 18, 21]
    for j, t in enumerate(times):
        ws.cell(row=6, column=2 + j, value=t)
    ws.cell(row=6, column=10, value="중위")
    ws.cell(row=6, column=11, value="최소")
    days = ["수요일, 4월 15일", "목요일, 4월 16일", "금요일, 4월 17일",
            "토요일, 4월 18일", "일요일, 4월 19일", "월요일, 4월 20일",
            "화요일, 4월 21일"]
    medians = [1010, 1015, 1013, 1015, 1015, 1014, 1014]
    mins = [1009, 1014, 1012, 1013, 1013, 1012, 1013]
    for i, (d, med, mn) in enumerate(zip(days, medians, mins)):
        r = 7 + i
        ws.cell(row=r, column=1, value=d)
        for j, t in enumerate(times):
            ws.cell(row=r, column=2 + j, value=med)  # 단순화
        ws.cell(row=r, column=10, value=med)
        ws.cell(row=r, column=11, value=mn)
    # Tempereture 섹션
    ws["A19"] = "Tempereture"
    for j, t in enumerate(times):
        ws.cell(row=20, column=2 + j, value=t)
    ws.cell(row=20, column=10, value="중위")
    ws.cell(row=20, column=11, value="취약")
    for i, (d, tm) in enumerate(zip(days, [20, 18, 17, 19, 21, 20, 22])):
        r = 21 + i
        ws.cell(row=r, column=1, value=d)
        ws.cell(row=r, column=10, value=tm)
    wb.save(path)
    return path, days, medians


def test_parse_and_applied_pressure(tmp_path):
    path, days, medians = _make_excel3_1(str(tmp_path / "w.xlsx"))
    fc = load_excel3_1(path)
    assert fc.capture.startswith("2026-04-15")
    assert fc.update_time == "17:00"          # 'Update time :' 값 파싱
    assert fc.days == days
    assert fc.times == [0, 3, 6, 9, 12, 15, 18, 21]
    assert fc.pressure_median[days[0]] == pytest.approx(1010)
    assert fc.temp_median[days[0]] == pytest.approx(20)

    # 적용 대기압 = 3~7일차 중위 평균 − 8 (엑셀3 M2 = AVERAGE(Y8:Y12)-8 과 동일)
    from wirye_capacity.weather import bid_days
    assert bid_days(fc) == days[2:7]
    mean_med = sum(medians[2:7]) / 5
    assert applied_pressure(fc) == pytest.approx(mean_med + C.WEATHER_SITE_OFFSET)
    assert applied_pressure(fc) == pytest.approx(1006.2)
    # 테스트일(D+0)과 다음날(D+1)은 쓰지 않는다 — 그 값이 바뀌어도 결과가 같아야 한다
    fc.pressure_median[days[0]] = 900.0
    fc.pressure_median[days[1]] = 900.0
    assert applied_pressure(fc) == pytest.approx(1006.2)


def test_site_offset_is_minus_8():
    assert C.WEATHER_SITE_OFFSET == -8.0


def test_missing_median_header_does_not_crash(tmp_path):
    """B1 회귀: 'Pressure' 섹션은 있으나 '중위' 헤더가 없어도 크래시하지 않음."""
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws["A1"] = "2026-04-15 18:09:12"
    ws["A5"] = "Pressure"
    # 시간 헤더만, '중위' 라벨 없음
    for j, t in enumerate([0, 3, 6, 9]):
        ws.cell(row=6, column=2 + j, value=t)
    ws.cell(row=7, column=1, value="수요일")
    p = str(tmp_path / "no_median.xlsx"); wb.save(p)
    fc = load_excel3_1(p)            # 예외 없이 반환되어야 함
    assert fc.pressure_median == {}  # 헤더 없으면 빈 dict


def test_rule_matches_excel3_template_formula():
    """우리 계산이 엑셀3 템플릿의 M2 수식과 같은 창을 쓰는가.

    엑셀3 '온도 Profile'!M2 = =AVERAGE(Y8:Y12)-8 이다. 날씨 블록은 6~12행(7일),
    Y열이 중위이므로 Y8:Y12 = 3~7일차 = D+2~D+6 다. 이 테스트는 **템플릿 수식을
    직접 읽어** 우리 상수와 대조한다 — 양식이 바뀌거나 누가 규칙을 되돌리면 깨진다.

    왜 중요한가: M2 는 어떤 셀도 참조하지 않는 표시용이다. 우리가 다른 창으로
    계산하면 파일이 '표시된 적용 대기압'과 '내용의 계산 기준'이 어긋난 채로
    나간다. 2026-08-25 까지 실제로 그 상태였다(우리 7일 평균 vs 엑셀 5일 평균).
    """
    import re

    from openpyxl import load_workbook

    from wirye_capacity.profile import DEFAULT_TEMPLATE, _WX

    ws = load_workbook(DEFAULT_TEMPLATE, data_only=False)[_WX["sheet"]]
    f = ws["M2"].value
    m = re.fullmatch(r"=AVERAGE\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)([+-]\d+(?:\.\d+)?)", f or "")
    assert m, f"M2 수식 형태가 바뀌었습니다: {f!r}"
    col_a, r1, col_b, r2, off = m.group(1), int(m.group(2)), m.group(3), int(m.group(4)), float(m.group(5))

    # 열은 중위 대기압 열(Y)이어야 한다
    from openpyxl.utils import get_column_letter
    assert col_a == col_b == get_column_letter(_WX["median_col"]), \
        f"M2 가 중위 열이 아닌 {col_a} 를 평균한다"
    # 행 범위 → 0-based 일차 슬라이스로 환산해 상수와 대조
    lo = r1 - _WX["first_row"]
    hi = r2 - _WX["first_row"] + 1
    assert (lo, hi) == C.WEATHER_BID_SLICE, \
        f"M2 는 {lo + 1}~{hi}일차를 평균하는데 상수는 {C.WEATHER_BID_SLICE} 다"
    assert off == C.WEATHER_SITE_OFFSET, f"M2 위치보정 {off} vs 상수 {C.WEATHER_SITE_OFFSET}"


def test_generated_file_pressure_matches_its_own_m2(tmp_path):
    """생성된 입찰파일의 M2 수식을 직접 계산해 보면 우리 적용 대기압과 같은가."""
    from openpyxl import load_workbook

    from wirye_capacity.pipeline import run_pipeline
    from wirye_capacity.profile import _WX
    from wirye_capacity.store import MeasurementStore
    from wirye_capacity.weather import WeatherForecast

    days = [f"D{i}" for i in range(7)]
    med = [1002.0, 999.1, 997.4, 1004.8, 1008.2, 1006.0, 1001.5]
    fc = WeatherForecast(capture="2026-07-07 18:00:00", days=days,
                         times=[0, 3, 6, 9, 12, 15, 18, 21],
                         pressure_median=dict(zip(days, med)))
    out = str(tmp_path / "bid.xlsx")
    store = MeasurementStore(":memory:")
    store.seed()
    res = run_pipeline(date="2026-07-07", store=store, output_path=out, forecast=fc)

    ws = load_workbook(out, data_only=False)[_WX["sheet"]]
    col = _WX["median_col"]
    r1, r2 = (_WX["first_row"] + C.WEATHER_BID_SLICE[0],
              _WX["first_row"] + C.WEATHER_BID_SLICE[1] - 1)
    ys = [ws.cell(row=r, column=col).value for r in range(r1, r2 + 1)]
    assert all(isinstance(v, (int, float)) for v in ys), f"Y열에 예보가 안 채워졌다: {ys}"
    excel_m2 = sum(ys) / len(ys) + C.WEATHER_SITE_OFFSET
    assert res.applied_pressure == pytest.approx(excel_m2), (
        f"파일이 표시할 값 {excel_m2:.3f} vs 실제 계산 기준 {res.applied_pressure:.3f}")
