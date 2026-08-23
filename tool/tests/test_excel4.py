"""엑셀4 실측데이터 시트 탐색 — 현장 파일의 실제 실패 형태 회귀 테스트.

현장 실패: 첫 시트가 '온도 Profile' 이어서 폴백이 엉뚱한 시트를 고르고,
헤더가 병합셀로 2행에 걸쳐('복수기압'/'실측') 있어 인식되지 않았다.
"""
from datetime import date

import pytest
from openpyxl import Workbook

from wirye_capacity.excel4 import load_excel4_records
from wirye_capacity.excel_io import ExcelOpenError
from wirye_capacity.store import MeasurementStore


def _make(path, meas_sheet="Test 실측 기록"):
    """실측 시트 앞에 함정 시트를 두고, 헤더를 병합셀 2행으로 구성."""
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "온도 Profile"                      # 이름 후보에 없는 첫 시트(함정)
    ws0.append(["온도", "CC이론", "CC현실화"])
    for t in range(-20, 41):
        ws0.append([t, 400 + t, 395 + t])
    ws = wb.create_sheet(meas_sheet)
    ws.append([None, None, None, None, "복수기압", None, "CC실측", None])   # 병합 상단
    ws.append(["날짜", "온도(℃)", "대기압(mbar)", "RH(%)", "실측", "설계", "(MW)", "W값"])
    ws.append([date(2026, 1, 6), 3.9, 1011.2, 23.0, 47.9, 31.1, 460.87, 4])
    ws.append(["2026.02.25", 15.1, 1011.6, 0.1, 48.9, 45.1, 445.60, 4])   # 문자열 날짜
    ws.append([date(2025, 9, 23), 25.1, 1007.7, 44.0, 59.4, 64.6, 414.54, 6])
    ws.append([None] * 8)                                                  # 빈 행
    wb.save(path)
    return path


def test_autodetect_skips_decoy_sheet_and_reads_merged_header(tmp_path):
    recs = load_excel4_records(_make(tmp_path / "e4.xlsx"))
    assert len(recs) == 3
    assert {r["sheet"] for r in recs} == {"Test 실측 기록"}      # 함정 시트 아님
    assert [r["date"] for r in recs] == ["2026-01-06", "2026-02-25", "2025-09-23"]
    assert recs[0]["cit"] == 3.9 and recs[0]["cc_meas"] == 460.87


def test_named_sheet_candidate_still_works(tmp_path):
    recs = load_excel4_records(_make(tmp_path / "e4b.xlsx", meas_sheet="실측데이터"))
    assert len(recs) == 3 and {r["sheet"] for r in recs} == {"실측데이터"}


def test_explicit_sheet_override(tmp_path):
    recs = load_excel4_records(_make(tmp_path / "e4c.xlsx"), sheet="Test 실측 기록")
    assert len(recs) == 3


def test_bad_sheet_name_lists_available_sheets(tmp_path):
    with pytest.raises(ExcelOpenError) as e:
        load_excel4_records(_make(tmp_path / "e4d.xlsx"), sheet="없는시트")
    assert "온도 Profile" in str(e.value) and "Test 실측 기록" in str(e.value)


def test_no_measurement_sheet_gives_diagnosis(tmp_path):
    wb = Workbook()
    wb.active.title = "표지"
    wb.active.append(["제목", "값"])
    p = tmp_path / "e4e.xlsx"
    wb.save(p)
    with pytest.raises(ExcelOpenError) as e:
        load_excel4_records(p)
    assert "--sheet" in str(e.value) and "표지" in str(e.value)


def test_backfill_fills_matching_seed_records(tmp_path):
    recs = load_excel4_records(_make(tmp_path / "e4f.xlsx"))
    store = MeasurementStore(tmp_path / "e4.db")
    store.seed()
    # 씨앗은 2026-08 정정으로 실제 날짜를 갖고 있다. 백필 시험이므로 비운다.
    store.conn.execute("UPDATE measurements SET date=NULL")
    store.conn.commit()
    assert store.count() == 31
    assert store.backfill_dates(recs) == 3           # 시드 3건과 (CIT, CC실측) 매칭
    dated = {r["date"] for r in store.list_up() if r.get("date")}
    assert dated == {"2026-01-06", "2026-02-25", "2025-09-23"}
    assert store.backfill_dates(recs) == 0           # 재실행 멱등
    store.close()
