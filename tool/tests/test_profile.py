"""Profile 생성(Phase 3) 검증 — 엑셀4 '온도 Profile' 셀값을 ±0.5MW로 재현."""
from pathlib import Path

import pytest

from wirye_capacity import constants as C
from wirye_capacity.profile import build_profile, write_xlsx
from wirye_capacity.store import MeasurementStore
from wirye_capacity.theory import TheoryEngine

TOL = 0.5  # 시운전 허용오차 (MW)


@pytest.fixture
def profile():
    eng = TheoryEngine()
    store = MeasurementStore(":memory:")
    store.seed()
    rows = build_profile(eng, store.correction_table(), pressure=1013, deg=1.028)
    store.close()
    return {r.temp: r for r in rows}


def test_covers_full_range(profile):
    assert min(profile) == -20 and max(profile) == 40
    assert len(profile) == 61


def test_cc_theory_matches_excel4_profile(profile):
    # 엑셀4 Profile D열 (= base + W)
    assert profile[-20].cc_theory == pytest.approx(444.567007, abs=TOL)
    assert profile[24].cc_theory == pytest.approx(419.25732, abs=TOL)
    assert profile[25].cc_theory == pytest.approx(418.292902, abs=TOL)


def test_cc_realized_gross_matches_excel4(profile):
    # 엑셀4 Profile G열: MIN(D + 보정값구간, cap)
    # 24°C: 419.257 + 2.62(20~25구간) = 421.877
    assert profile[24].cc_real_gross == pytest.approx(421.877, abs=TOL)
    # 25°C: 418.293 + (−2.39)(25~30구간) = 415.903
    assert profile[25].cc_real_gross == pytest.approx(415.903, abs=TOL)
    # −6°C: 이론 + 8.78(−14~0 고정)
    assert profile[-6].cc_real_gross == pytest.approx(profile[-6].cc_theory + 8.78, abs=1e-6)


def test_realized_net_is_gross_minus_aux(profile):
    r = profile[25]
    assert r.cc_real_net == pytest.approx(r.cc_real_gross - C.CC_AUX, abs=1e-6)
    # 모든 입찰값이 상한 이내
    assert all(p.cc_real_net <= C.BID_CAP_NET + 1e-9 for p in profile.values())


def test_gt_theory_close_and_st_consistency(profile):
    # GT 이론 = base_gt×k + W·0.657 (엑셀4 B50=275.440)
    assert profile[24].gt_theory == pytest.approx(275.440, abs=TOL)
    # ST = CC − GT 정합 (전 온도)
    for p in profile.values():
        assert p.st_theory == pytest.approx(p.cc_theory - p.gt_theory, abs=1e-9)
        assert p.st_real == pytest.approx(p.cc_real_gross - p.gt_real, abs=1e-9)


def test_shaft_limit_no_correction(profile):
    # −20~−14°C: 보정값 0 (이론값 고정)
    assert profile[-18].correction == 0.0
    assert profile[-18].cc_real_gross == pytest.approx(profile[-18].cc_theory, abs=1e-9)


def test_write_xlsx_roundtrip(tmp_path, profile):
    import openpyxl
    out = write_xlsx(list(profile.values()), str(tmp_path / "profile.xlsx"))
    assert Path(out).exists()
    wb = openpyxl.load_workbook(out)
    ws = wb["온도 Profile"]
    # 헤더 다음 첫 데이터행(−20°C) CC 이론 확인
    # 컬럼: 1温 2GT이론 3ST이론 4CC이론 ...
    first_temp = ws.cell(row=5, column=1).value
    assert first_temp == -20
    assert ws.cell(row=5, column=4).value == pytest.approx(444.57, abs=TOL)
