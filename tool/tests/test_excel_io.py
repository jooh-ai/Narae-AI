"""엑셀 열기 오류 처리 + 엑셀4 날짜 백필 검증."""
import pytest

from wirye_capacity.excel_io import ExcelOpenError, load_workbook_safe


def test_bad_zip_gives_friendly_error(tmp_path):
    """DRM/손상 파일(zip 아님) → 원인·해결책 담은 ExcelOpenError."""
    p = tmp_path / "secured.xlsx"
    p.write_bytes(b"not a zip at all")
    with pytest.raises(ExcelOpenError) as ei:
        load_workbook_safe(p)
    msg = str(ei.value)
    assert "보안" in msg and "다른 이름으로 저장" in msg
    assert "secured.xlsx" in msg


def test_missing_file_error(tmp_path):
    with pytest.raises(ExcelOpenError, match="파일이 없습니다"):
        load_workbook_safe(tmp_path / "none.xlsx")


def test_valid_file_opens(tmp_path):
    from openpyxl import Workbook
    p = tmp_path / "ok.xlsx"
    Workbook().save(p)
    assert load_workbook_safe(p) is not None


def _make_excel4(path, rows):
    """엑셀4 '실측데이터' 헤더 레이아웃 합성."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "실측데이터"
    ws.append(["위례열병합발전소  IGV Turn up 실측 데이터"])       # 제목 행
    ws.append(["날짜", "온도(°C)", "대기압(mbar)", "RH(%)", "복수기압실측(mbar)",
               "복수기설계(mbar)", "CC 실측(MW)", "W값(IGV)", "이론기준값(MW)",
               "보정값(MW)", "비고"])
    for d, cit, cc in rows:
        ws.append([d, cit, 1011.2, 23.0, 47.9, 31.1, cc, 4, 451.16, 5.713, "겨울"])
    wb.save(path)
    return path


def test_backfill_dates_from_excel4(tmp_path):
    """시드(날짜 없음) → 엑셀4 날짜가 (CC실측·CIT) 매칭으로 채워진다."""
    from wirye_capacity.excel4 import load_excel4_records
    from wirye_capacity.store import MeasurementStore

    s = MeasurementStore(":memory:")
    s.seed()
    assert all(not r.get("date") for r in s.list_up())        # 시드는 날짜 없음
    # 시드에 실제 존재하는 (cit, cc_meas) 2건으로 엑셀4 합성
    p = _make_excel4(str(tmp_path / "e4.xlsx"),
                     [("2025-01-06", 3.9, 460.87), ("2025-02-10", 1.9, 466.93)])
    recs = load_excel4_records(p)
    assert len(recs) == 2 and recs[0]["date"] == "2025-01-06"

    n = s.backfill_dates(recs)
    assert n == 2
    dates = {r["date"] for r in s.list_up() if r.get("date")}
    assert dates == {"2025-01-06", "2025-02-10"}
    # 재실행해도 중복 배정 없음(이미 채워진 레코드는 대상 아님)
    assert s.backfill_dates(recs) == 0
    s.close()


def test_backfill_ignores_unmatched(tmp_path):
    """엑셀4에 없는 값이면 채우지 않는다(오배정 방지)."""
    from wirye_capacity.excel4 import load_excel4_records
    from wirye_capacity.store import MeasurementStore
    s = MeasurementStore(":memory:")
    s.seed()
    p = _make_excel4(str(tmp_path / "e4.xlsx"), [("2025-01-06", 99.9, 999.99)])
    assert s.backfill_dates(load_excel4_records(p)) == 0
    s.close()


def test_excel4_date_formats(tmp_path):
    """날짜 셀이 datetime 이거나 '2025.01.06' 형식이어도 파싱."""
    from datetime import datetime

    from wirye_capacity.excel4 import load_excel4_records
    p = _make_excel4(str(tmp_path / "e4.xlsx"),
                     [(datetime(2025, 1, 6), 3.9, 460.87), ("2025.02.10", 1.9, 466.93)])
    recs = load_excel4_records(p)
    assert [r["date"] for r in recs] == ["2025-01-06", "2025-02-10"]
