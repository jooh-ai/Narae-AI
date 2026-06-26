"""엑셀3 템플릿 채우기 검증 — Mode3!B5:C65에 현실화 GT/ST 기입, 양식·수식 보존."""
import pytest

from wirye_capacity import constants as C
from wirye_capacity.profile import (
    DEFAULT_TEMPLATE, build_profile, fill_excel3_template,
)
from wirye_capacity.store import MeasurementStore
from wirye_capacity.theory import TheoryEngine
from wirye_capacity.weather import WeatherForecast


@pytest.fixture
def ctx():
    eng = TheoryEngine()
    s = MeasurementStore(":memory:")
    s.seed()
    table = s.correction_table()
    s.close()
    return eng, table


def test_template_exists():
    assert DEFAULT_TEMPLATE.exists()


def test_fills_mode3_input_and_preserves_format(tmp_path, ctx):
    import openpyxl
    eng, table = ctx
    out = fill_excel3_template(str(tmp_path / "bid.xlsx"), engine=eng,
                               correction_table=table, pressure=1013, deg=1.028)
    rows = {r.temp: r for r in build_profile(eng, table, pressure=1013, deg=1.028)}

    wb = openpyxl.load_workbook(out)
    # 양식 보존: 시트·수식 그대로
    assert wb.sheetnames == ["온도 Profile", "Mode3", "Mode5", "Mode1", "열전비 Sheet", "Sheet1"]
    assert wb["온도 Profile"]["C6"].value == "=Mode3!G5"      # 파생 수식 유지

    m3 = wb["Mode3"]
    # 행5 = −20°C: B=GT현실화, C=ST현실화, B+C = CC현실화 Gross
    assert m3["A5"].value == -20
    assert m3["B5"].value == pytest.approx(rows[-20].gt_real, abs=1e-6)
    assert m3["C5"].value == pytest.approx(rows[-20].st_real, abs=1e-6)
    assert m3["B5"].value + m3["C5"].value == pytest.approx(rows[-20].cc_real_gross, abs=1e-6)
    # 행65 = 40°C
    assert m3["A65"].value == 40
    assert m3["B65"].value + m3["C65"].value == pytest.approx(rows[40].cc_real_gross, abs=1e-6)


def test_pressure_baked_into_values(tmp_path, ctx):
    import openpyxl
    eng, table = ctx
    out_a = fill_excel3_template(str(tmp_path / "a.xlsx"), engine=eng,
                                 correction_table=table, pressure=1013, deg=1.028)
    out_b = fill_excel3_template(str(tmp_path / "b.xlsx"), engine=eng,
                                 correction_table=table, pressure=1000, deg=1.028)
    b_a = openpyxl.load_workbook(out_a)["Mode3"]["B30"].value
    b_b = openpyxl.load_workbook(out_b)["Mode3"]["B30"].value
    # 대기압↓ → 출력↓ (값에 대기압이 반영됨)
    assert b_b < b_a


def test_weather_block_filled(tmp_path, ctx):
    import openpyxl
    eng, table = ctx
    fc = WeatherForecast(
        capture="2026-04-15 18:09:12",
        days=["수요일, 4월 15일", "목요일, 4월 16일"],
        times=[0, 3, 6, 9, 12, 15, 18, 21],
        pressure_median={"수요일, 4월 15일": 1010.0, "목요일, 4월 16일": 1015.0},
        pressure_min={"수요일, 4월 15일": 1009.0, "목요일, 4월 16일": 1014.0},
        pressure_grid={"수요일, 4월 15일": [1010, 1009, 1010, 1010, 1010, 1009, 1009, 1011],
                       "목요일, 4월 16일": [1014, 1015, 1015, 1017, 1017, 1015, 1015, 1015]})
    out = fill_excel3_template(str(tmp_path / "bid.xlsx"), engine=eng,
                               correction_table=table, pressure=1006, deg=1.028, forecast=fc)
    ws = openpyxl.load_workbook(out)["온도 Profile"]
    assert ws.cell(row=6, column=16).value == "수요일, 4월 15일"   # P6 일자
    assert ws.cell(row=6, column=17).value == 1010                # Q6 0시
    assert ws.cell(row=6, column=25).value == pytest.approx(1010) # Y6 중위
