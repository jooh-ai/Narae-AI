"""검증 모드(Phase 5) 테스트 — 단계별 대조·허용오차 판정·기준 .xlsx 읽기."""
import pytest

from wirye_capacity.profile import build_profile, write_xlsx
from wirye_capacity.store import MeasurementStore
from wirye_capacity.theory import TheoryEngine
from wirye_capacity.verify import (
    compare_profile, read_reference_xlsx, ref_from_profile,
)


@pytest.fixture
def rows():
    eng = TheoryEngine()
    s = MeasurementStore(":memory:")
    s.seed()
    rws = build_profile(eng, s.correction_table(), pressure=1013, deg=1.028)
    s.close()
    return rws


def test_self_compare_passes(rows):
    ref = ref_from_profile(rows)
    rep = compare_profile(rows, ref, tol=0.5)
    assert rep.passed
    assert rep.failures == []
    assert rep.max_abs_diff == pytest.approx(0.0)


def test_perturbed_reference_fails(rows):
    ref = ref_from_profile(rows)
    ref[25]["cc_real_gross"] += 1.0   # 1 MW 어긋남 → 허용 0.5 초과
    rep = compare_profile(rows, ref, tol=0.5)
    assert not rep.passed
    bad = rep.failures
    assert any(r["temp"] == 25 and r["field"] == "cc_real_gross" for r in bad)
    assert rep.max_abs_diff == pytest.approx(1.0, abs=1e-9)


def test_within_tolerance_passes(rows):
    ref = ref_from_profile(rows)
    ref[20]["cc_theory"] += 0.3   # 0.3 MW < 0.5 → 합격
    rep = compare_profile(rows, ref, tol=0.5)
    assert rep.passed


def test_roundtrip_via_tool_xlsx(tmp_path, rows):
    path = write_xlsx(rows, str(tmp_path / "p.xlsx"))
    ref = read_reference_xlsx(path, layout="tool")
    rep = compare_profile(rows, ref, tol=0.5)
    assert rep.passed            # 반올림(0.01) 차이뿐 → 합격
    assert rep.max_abs_diff < 0.01


def test_read_excel4_layout(tmp_path, rows):
    # 엑셀4 'Mode3' 형식(A온도 row6~, D CC이론, G CC현실화) 합성 후 읽기
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for i, r in enumerate(rows):
        row = 6 + i
        ws.cell(row=row, column=1, value=r.temp)
        ws.cell(row=row, column=4, value=round(r.cc_theory, 2))
        ws.cell(row=row, column=7, value=round(r.cc_real_gross, 2))
    p = str(tmp_path / "excel4.xlsx")
    wb.save(p)
    ref = read_reference_xlsx(p, layout="excel4")
    assert -20 in ref and 40 in ref
    rep = compare_profile(rows, ref, tol=0.5, fields=["cc_theory", "cc_real_gross"])
    assert rep.passed
