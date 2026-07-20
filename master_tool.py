# -*- coding: utf-8 -*-
"""
자재마스터 통합 도구 (MasterTool)
================================================

세 가지 기능을 하나의 대화형 메뉴로 통합한 도구입니다.

  1) 사양 분배      : M열 [라벨:값]…  →  O~AH 개별속성 칸에 '값'만 분배
  2) Working Q'ty   : 재고 R열(max)  →  자재마스터 N열에 기입
  3) 마스터 병합    : 수정본(제어파트)  →  원본에 A열 기준 반영

공통 특성
---------
- 대상 시트는 이름에 '위례'가 들어간 시트를 자동 선택(없으면 첫 시트).
- 원본 파일은 건드리지 않고, 결과는 새 파일로 저장합니다.
- 셀 서식(폰트·색·테두리·열 너비 등)은 그대로 보존합니다.
- 새로 기입/변경한 셀만 빨간 글자색으로 강조합니다.

실행 방법
---------
[exe로 쓰는 경우]   MasterTool.exe 를 더블클릭 → 메뉴에서 번호 선택
                    처리할 엑셀 파일들을 exe와 같은 폴더에 두면 목록에서 고를 수 있습니다.
[파이썬으로 쓰는 경우]
    1) 한 번만:  pip install openpyxl
    2) 이 파일을 실행 (VSCode F5 또는  python master_tool.py)

열 위치가 표준과 다르면 아래 '기본 열 설정' 상수를 수정하세요.
"""

import os
import sys
import re
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import column_index_from_string
except ImportError:
    print("✗ openpyxl 라이브러리가 필요합니다. 터미널에서 다음을 실행하세요:")
    print("    pip install openpyxl")
    input("Enter 키를 누르면 종료합니다...")
    sys.exit(1)


# ═══════════════════════════════════════════════
#                기본 열 설정 (표준 마스터)
# ═══════════════════════════════════════════════
HEADER_ROW = 1            # 헤더 행 (이 행 다음부터 데이터)

# [1] 사양 분배
SPEC_COL = "M"            # 사양 컬럼
SPEC_PN_COL = "L"         # MAKER P/N (사양 없는 행 보완용) 컬럼
ATTR_START_COL = "O"      # 개별속성 시작열
ATTR_END_COL = "AH"       # 개별속성 끝열

# [2] Working Q'ty
WQ_MASTER_KEY = "A"       # 자재마스터 자재(키) 열
WQ_TARGET = "N"           # 자재마스터 Working Q'ty (기입 대상) 열
WQ_INV_KEY = "A"          # 재고 자재(키) 열
WQ_INV_VAL = "R"          # 재고 max (가져올 값) 열

# [3] 마스터 병합
MG_KEY = "A"              # 자재(키) 열
MG_VALUE_COLS = ["B", "H", "J", "K", "L", "N"]  # 값 있을 때만 덮어쓰는 열
MG_MIR_START = "O"        # 완전 미러 시작열 (빈칸도 덮어씀)
MG_MIR_END = "AH"         # 완전 미러 끝열
# ═══════════════════════════════════════════════

RED_ARGB = "FFFF0000"
MAKER_PN_NORM = "MAKERPN"


# ─────────────────────────────────────────────
# 공통 헬퍼
# ─────────────────────────────────────────────
def val_text(v):
    """값을 문자열로 (None → '')."""
    return "" if v is None else str(v)


def cell_text(cell):
    return val_text(cell.value)


def is_empty(v):
    """공백/제로폭 문자만 있으면 빈 값으로 간주."""
    if v is None:
        return True
    return re.sub(r"[\s ​‌‍﻿]+", "", str(v)) == ""


def norm_label(s):
    """속성명 정규화: ':' 앞부분 대문자화 + 공백·점·슬래시 제거.
    'MAKER P/N' → 'MAKERPN'."""
    t = "" if s is None else str(s)
    m = re.search(r"[:：]", t)
    if m:
        t = t[: m.start()]
    return re.sub(r"[\s.\/]", "", t.upper())


def norm_value(s):
    """매칭/중복 비교용 값 정규화 (공백·언더스코어 제거, 소문자)."""
    return re.sub(r"[\s_]", "", ("" if s is None else str(s)).lower())


def key_norm(v):
    """자재(키) 비교용 정규화. 정수형 실수(12345.0)는 '12345'로 통일."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def to_number(v):
    """숫자로 변환. 변환 불가 시 None."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def same_value(a, b):
    """두 셀 값이 사실상 같은지 (사용자 시각)."""
    return val_text(a).strip() == val_text(b).strip()


def apply_red_font(cell):
    """기존 폰트 속성을 보존하면서 글자색만 빨강으로."""
    f = cell.font
    cell.font = Font(
        name=f.name, size=f.size, bold=f.bold, italic=f.italic,
        vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
        color=RED_ARGB,
    )


def write_cell(cell, value, highlight):
    cell.value = value
    if highlight:
        apply_red_font(cell)


def resolve_path(base_dir, path):
    return path if os.path.isabs(path) else os.path.join(base_dir, path)


def pick_sheet(wb, preferred=""):
    """preferred 지정 시 그 시트, 아니면 '위례' 포함 시트, 없으면 첫 시트."""
    if preferred and preferred in wb.sheetnames:
        return wb[preferred]
    chosen = next((n for n in wb.sheetnames if "위례" in n), wb.sheetnames[0])
    return wb[chosen]


def out_path_for(base_dir, in_path, suffix, keep_vba):
    base = os.path.splitext(os.path.basename(in_path))[0]
    stamp = datetime.now().strftime("%y%m%d_%H%M")
    ext = ".xlsm" if keep_vba else ".xlsx"
    tag = f"_{suffix}" if suffix else ""
    return os.path.join(base_dir, f"{base}{tag}_{stamp}{ext}")


# ─────────────────────────────────────────────
# 사양 파싱 (기능 1)
# ─────────────────────────────────────────────
def push_segment(content, entries, unlabeled):
    t = (content or "").strip()
    if not t:
        return
    m = re.search(r"[:：]", t)
    if m:
        label = t[: m.start()].strip()
        value = t[m.start() + 1:].strip()
        if label:
            entries.append({"label": label, "value": value, "raw": t})
            return
    unlabeled.append(t)


def parse_spec(spec):
    """M열 사양 문자열 파싱 → (entries, unlabeled)."""
    entries, unlabeled = [], []
    text = "" if spec is None else str(spec)
    for m in re.finditer(r"\[([^\[\]]*)\]", text):
        push_segment(m.group(1), entries, unlabeled)
    remainder = re.sub(r"\[[^\[\]]*\]", " ", text).strip()
    remainder = re.sub(r"^[\s_]+|[\s_]+$", "", remainder)
    if remainder:
        known = set()
        for e in entries:
            if e["value"]:
                known.add(norm_value(e["value"]))
            known.add(norm_value(e["raw"]))
        for u in unlabeled:
            known.add(norm_value(u))
        if norm_value(remainder) not in known:
            push_segment(remainder, entries, unlabeled)
    return entries, unlabeled


# ─────────────────────────────────────────────
# 기능 1: 사양 분배
# ─────────────────────────────────────────────
def op_spec(base_dir, in_path, highlight_red, pn_fallback):
    keep_vba = in_path.lower().endswith(".xlsm")
    print(f"▶ 파일 로드 중: {in_path}")
    wb = load_workbook(in_path, keep_vba=keep_vba)
    ws = pick_sheet(wb)
    print(f"  · 대상 시트: '{ws.title}'")

    spec_col = column_index_from_string(SPEC_COL)
    pn_col = column_index_from_string(SPEC_PN_COL)
    attr_start = column_index_from_string(ATTR_START_COL)
    attr_end = column_index_from_string(ATTR_END_COL)

    max_row = ws.max_row
    total_rows = max_row - HEADER_ROW
    print(f"  · 데이터 {total_rows}행, 개별속성 {ATTR_START_COL}~{ATTR_END_COL}열")

    rows_with_spec = value_cells = cleared_cells = etc_rows = 0
    pn_fallback_rows = no_etc_warn = specless_cleaned = specless_skipped = 0
    sample_unmatched, sample_skipped = [], []

    for r in range(HEADER_ROW + 1, max_row + 1):
        spec_val = cell_text(ws.cell(row=r, column=spec_col))
        entries, unlabeled = parse_spec(spec_val)

        # (A) 사양에 [라벨:값] 항목이 없는 행 → MAKER P/N만 남기고 정리
        if not entries:
            pn_l_raw = cell_text(ws.cell(row=r, column=pn_col)).strip()
            pn_l = norm_value(pn_l_raw)
            nonempty, keep_col, keep_val, keep_is_fallback = [], None, None, False
            for c in range(attr_start, attr_end + 1):
                v = cell_text(ws.cell(row=r, column=c))
                if is_empty(v):
                    continue
                s = str(v)
                nonempty.append((c, s))
                if keep_col is not None:
                    continue
                ci = next((i for i, ch in enumerate(s) if ch in ":："), -1)
                vv = s[ci + 1:].strip() if ci >= 0 else ""
                if norm_label(s) == MAKER_PN_NORM:
                    keep_col = c
                    if vv:
                        keep_val = vv
                    elif pn_fallback and pn_l_raw:
                        keep_val, keep_is_fallback = pn_l_raw, True
                    else:
                        keep_val = None
                elif pn_l and norm_value(s) == pn_l:
                    keep_col = c
                    keep_val = vv if vv else s.strip()

            if keep_col is None:
                if nonempty:
                    specless_skipped += 1
                    if len(sample_skipped) < 10:
                        sample_skipped.append(f"{r}행")
                continue

            for c, s in nonempty:
                if c == keep_col:
                    if keep_val is None:
                        ws.cell(row=r, column=c).value = None
                        cleared_cells += 1
                    elif s.strip() != keep_val:
                        write_cell(ws.cell(row=r, column=c), keep_val, highlight_red)
                        value_cells += 1
                        if keep_is_fallback:
                            pn_fallback_rows += 1
                else:
                    ws.cell(row=r, column=c).value = None
                    cleared_cells += 1
            specless_cleaned += 1
            continue

        # 사양이 있는 행: 라벨→열 맵 + ETC 열
        label_to_col, etc_col, attr_cols = {}, 0, []
        for c in range(attr_start, attr_end + 1):
            txt = cell_text(ws.cell(row=r, column=c))
            if is_empty(txt):
                continue
            norm = norm_label(txt)
            if not norm:
                continue
            attr_cols.append(c)
            if norm not in label_to_col:
                label_to_col[norm] = c
            if "ETC" in norm:
                etc_col = c

        rows_with_spec += 1
        col_value, unmatched = {}, []
        for e in entries:
            norm = norm_label(e["label"])
            if "ETC" in norm:
                if e["value"]:
                    unmatched.append(e["value"])
                continue
            target = label_to_col.get(norm)
            if target is not None and target != etc_col:
                if e["value"]:
                    col_value[target] = e["value"]
            else:
                unmatched.append(e["raw"])
        for u in unlabeled:
            unmatched.append(u)
        if unmatched:
            if etc_col:
                col_value[etc_col] = ", ".join(unmatched)
                etc_rows += 1
            else:
                no_etc_warn += 1
                if len(sample_unmatched) < 10:
                    sample_unmatched.append(f"{SPEC_COL}{r}: {' | '.join(unmatched)}")

        for c in attr_cols:
            if c in col_value:
                write_cell(ws.cell(row=r, column=c), col_value[c], highlight_red)
                value_cells += 1
            else:
                ws.cell(row=r, column=c).value = None
                cleared_cells += 1

        if (r - HEADER_ROW) % 500 == 0:
            print(f"  ... {r - HEADER_ROW}/{total_rows} 행 처리")

    print(f"  → 처리 완료: 사양 보유 {rows_with_spec}행, 사양 없는 행 정리 {specless_cleaned}행, "
          f"값 기입 {value_cells}칸, 빈칸 정리 {cleared_cells}칸, ETC 모음 {etc_rows}행, L열 보완 {pn_fallback_rows}칸")
    if specless_skipped:
        print(f"  ⚠ MAKER P/N을 못 찾아 건드리지 않은 사양없는 행 {specless_skipped}건")
        for s in sample_skipped:
            print(f"      · {s}")
    if no_etc_warn:
        print(f"  ⚠ ETC 칸을 못 찾아 미매칭 데이터 기입 못한 행 {no_etc_warn}건")
        for s in sample_unmatched:
            print(f"      · {s}")

    out_path = out_path_for(base_dir, in_path, "사양분배", keep_vba)
    wb.save(out_path)
    print(f"✅ 저장 완료: {out_path}")


# ─────────────────────────────────────────────
# 기능 2: Working Q'ty 채우기
# ─────────────────────────────────────────────
def op_workqty(base_dir, master_path, inv_path, highlight_red, overwrite_existing):
    mkey = column_index_from_string(WQ_MASTER_KEY)
    mtgt = column_index_from_string(WQ_TARGET)
    ikey = column_index_from_string(WQ_INV_KEY)
    ival = column_index_from_string(WQ_INV_VAL)

    print(f"▶ 재고 파일 로드 중: {inv_path}")
    inv_wb = load_workbook(inv_path, data_only=True)
    inv_ws = pick_sheet(inv_wb)
    print(f"  · 재고 시트: '{inv_ws.title}'")

    inv_map, dup_keys, inv_rows, inv_skipped = {}, set(), 0, 0
    for r in range(HEADER_ROW + 1, inv_ws.max_row + 1):
        k = key_norm(inv_ws.cell(row=r, column=ikey).value)
        if k == "":
            continue
        raw = inv_ws.cell(row=r, column=ival).value
        num = to_number(raw)
        if num is None:
            inv_skipped += 1
            continue
        inv_rows += 1
        if k in inv_map:
            dup_keys.add(k)
            if num > inv_map[k][0]:
                inv_map[k] = (num, raw)
        else:
            inv_map[k] = (num, raw)
    print(f"  · 재고 자재 {len(inv_map)}종 인덱싱 (유효 {inv_rows}건, 중복 {len(dup_keys)}종, "
          f"숫자아님 {inv_skipped}건 제외)")

    print(f"▶ 자재마스터 로드 중: {master_path}")
    keep_vba = master_path.lower().endswith(".xlsm")
    wb = load_workbook(master_path, keep_vba=keep_vba)
    ws = pick_sheet(wb)
    print(f"  · 자재마스터 시트: '{ws.title}'")

    written = no_match = skipped_existing = empty_key = 0
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        k = key_norm(ws.cell(row=r, column=mkey).value)
        if k == "":
            empty_key += 1
            continue
        hit = inv_map.get(k)
        if hit is None:
            no_match += 1
            continue
        tgt = ws.cell(row=r, column=mtgt)
        if not overwrite_existing and not is_empty(tgt.value):
            skipped_existing += 1
            continue
        tgt.value = hit[1]
        if highlight_red:
            apply_red_font(tgt)
        written += 1
        if (r - HEADER_ROW) % 1000 == 0:
            print(f"  ... {r - HEADER_ROW}/{ws.max_row - HEADER_ROW} 행 처리")

    print(f"  → 완료: N열 기입 {written}행, 재고에 없음 {no_match}행, A열 비어있음 {empty_key}행" +
          ("" if overwrite_existing else f", 기존값 보존 {skipped_existing}행"))

    out_path = out_path_for(base_dir, master_path, "WorkingQty", keep_vba)
    wb.save(out_path)
    print(f"✅ 저장 완료: {out_path}")


# ─────────────────────────────────────────────
# 기능 3: 마스터 병합 (수정본 → 원본)
# ─────────────────────────────────────────────
def op_merge(base_dir, orig_path, mod_path, highlight_red):
    key_c = column_index_from_string(MG_KEY)
    value_cols = [column_index_from_string(c) for c in MG_VALUE_COLS]
    mir_start = column_index_from_string(MG_MIR_START)
    mir_end = column_index_from_string(MG_MIR_END)
    h_col = column_index_from_string("H")

    print(f"▶ 원본 로드 중: {orig_path}")
    keep_vba = orig_path.lower().endswith(".xlsm")
    owb = load_workbook(orig_path, keep_vba=keep_vba)
    ows = pick_sheet(owb)
    print(f"  · 원본 시트: '{ows.title}'")

    print(f"▶ 수정본 로드 중: {mod_path}")
    mwb = load_workbook(mod_path, data_only=True)
    mws = pick_sheet(mwb)
    print(f"  · 수정본 시트: '{mws.title}'")

    # 원본 A열 → 행 인덱스
    orig_index, dup_orig = {}, 0
    for r in range(HEADER_ROW + 1, ows.max_row + 1):
        k = key_norm(ows.cell(row=r, column=key_c).value)
        if k == "":
            continue
        if k in orig_index:
            dup_orig += 1
            continue
        orig_index[k] = r
    print(f"  · 원본 자재 {len(orig_index)}종 인덱싱" +
          (f" (중복 {dup_orig}건은 첫 행만 사용)" if dup_orig else ""))

    matched = not_found = empty_key = 0
    value_changed = mirror_changed = mirror_cleared = h_changed = 0
    sample_not_found, sample_h = [], []

    for mr in range(HEADER_ROW + 1, mws.max_row + 1):
        k = key_norm(mws.cell(row=mr, column=key_c).value)
        if k == "":
            empty_key += 1
            continue
        orow = orig_index.get(k)
        if orow is None:
            not_found += 1
            if len(sample_not_found) < 15:
                sample_not_found.append(k)
            continue
        matched += 1

        # 값 있을 때만 덮어쓰는 열 (B/H/J/K/L/N)
        for col in value_cols:
            mv = mws.cell(row=mr, column=col).value
            if is_empty(mv):
                continue
            ocell = ows.cell(row=orow, column=col)
            if not same_value(ocell.value, mv):
                old = ocell.value
                ocell.value = mv
                if highlight_red:
                    apply_red_font(ocell)
                value_changed += 1
                if col == h_col:
                    h_changed += 1
                    if len(sample_h) < 15:
                        sample_h.append(f"{k}: '{val_text(old)}'→'{val_text(mv)}' (원본 {orow}행)")

        # 빈칸도 덮어쓰는 열 (O~AH 완전 미러)
        for col in range(mir_start, mir_end + 1):
            mv = mws.cell(row=mr, column=col).value
            ocell = ows.cell(row=orow, column=col)
            if same_value(ocell.value, mv):
                continue
            if is_empty(mv):
                ocell.value = None
                mirror_cleared += 1
            else:
                ocell.value = mv
                if highlight_red:
                    apply_red_font(ocell)
                mirror_changed += 1

        if (mr - HEADER_ROW) % 500 == 0:
            print(f"  ... 수정본 {mr - HEADER_ROW}/{mws.max_row - HEADER_ROW} 행 처리")

    print(f"  → 병합 완료: 매칭·갱신 {matched}행, 원본에 없음 {not_found}행, A열 비어있음 {empty_key}행")
    print(f"     · B/H/J/K/L/N 변경 {value_changed}칸 (그중 H열 부서 변경 {h_changed}행)")
    print(f"     · O~AH 변경 {mirror_changed}칸 기입 + {mirror_cleared}칸 비움")
    if h_changed:
        print("     · 부서(H열) 변경 내역:")
        for s in sample_h:
            print(f"         - {s}")
    if not_found:
        print(f"  ⚠ 수정본에만 있고 원본에 없는 자재 {not_found}건 (반영 못함):")
        for s in sample_not_found:
            print(f"         - {s}")

    out_path = out_path_for(base_dir, orig_path, "merged", keep_vba)
    owb.save(out_path)
    print(f"✅ 저장 완료: {out_path}")


# ─────────────────────────────────────────────
# 대화형 입력 헬퍼
# ─────────────────────────────────────────────
def clean_path_input(s):
    """드래그&드롭/붙여넣기 경로 정리: 앞뒤 공백·따옴표 제거."""
    s = s.strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in "\"'":
        s = s[1:-1]
    return s.strip()


def list_excel_files(base_dir):
    try:
        names = os.listdir(base_dir)
    except OSError:
        return []
    files = [n for n in names
             if n.lower().endswith((".xlsx", ".xlsm")) and not n.startswith("~$")]
    files.sort()
    return files


def prompt_file(base_dir, label):
    """엑셀 파일 하나를 대화형으로 받는다. 목록 번호 선택 / 경로 입력 / 드래그&드롭."""
    files = list_excel_files(base_dir)
    while True:
        print(f"\n[{label}] 파일을 선택하세요.")
        if files:
            print("  (현재 폴더의 엑셀 파일)")
            for i, f in enumerate(files, 1):
                print(f"    {i}. {f}")
            print("  번호를 입력하거나, 파일을 창에 끌어다 놓거나, 경로를 직접 입력하세요.")
        else:
            print("  파일을 창에 끌어다 놓거나 경로를 직접 입력하세요.")
        raw = input("  > ").strip()
        if raw == "":
            print("  ✗ 입력이 비었습니다. 다시 시도하세요.")
            continue
        # 번호 선택
        if files and raw.isdigit():
            idx = int(raw)
            if 1 <= idx <= len(files):
                return os.path.join(base_dir, files[idx - 1])
            print("  ✗ 목록에 없는 번호입니다.")
            continue
        # 경로 입력
        p = clean_path_input(raw)
        p = resolve_path(base_dir, p)
        if os.path.exists(p):
            return p
        print(f"  ✗ 파일을 찾을 수 없습니다: {p}")


def prompt_yes_no(label, default=True):
    d = "Y/n" if default else "y/N"
    while True:
        raw = input(f"{label} ({d}) > ").strip().lower()
        if raw == "":
            return default
        if raw in ("y", "yes", "예", "ㅇ"):
            return True
        if raw in ("n", "no", "아니오", "ㄴ"):
            return False
        print("  ✗ y 또는 n 으로 답하세요.")


# ─────────────────────────────────────────────
# 메뉴
# ─────────────────────────────────────────────
BANNER = """
╔══════════════════════════════════════════════╗
║            자재마스터 통합 도구                ║
╠══════════════════════════════════════════════╣
║  1) 사양 분배      (M열 → O~AH 개별속성)       ║
║  2) Working Q'ty   (재고 R → 자재마스터 N)     ║
║  3) 마스터 병합    (수정본 → 원본, A열 기준)   ║
║  0) 종료                                       ║
╚══════════════════════════════════════════════╝"""


def get_base_dir():
    """exe(frozen)면 exe가 있는 폴더, 아니면 스크립트 폴더/현재 폴더."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    try:
        return os.path.dirname(os.path.abspath(__file__))
    except NameError:
        return os.getcwd()


def run_menu_once(base_dir, choice):
    if choice == "1":
        print("\n── [1] 사양 분배 ──")
        in_path = prompt_file(base_dir, "자재마스터(사양 포함)")
        pn_fb = prompt_yes_no("사양 없는 행을 L열(MAKER P/N) 값으로 보완할까요?", default=False)
        hl = prompt_yes_no("변경 셀을 빨강으로 표시할까요?", default=True)
        op_spec(base_dir, in_path, hl, pn_fb)
    elif choice == "2":
        print("\n── [2] Working Q'ty 채우기 ──")
        master = prompt_file(base_dir, "자재마스터 (N열 채울 파일)")
        inv = prompt_file(base_dir, "재고 파일")
        ow = prompt_yes_no("N열에 기존 값이 있어도 덮어쓸까요?", default=True)
        hl = prompt_yes_no("변경 셀을 빨강으로 표시할까요?", default=True)
        op_workqty(base_dir, master, inv, hl, ow)
    elif choice == "3":
        print("\n── [3] 마스터 병합 ──")
        orig = prompt_file(base_dir, "원본 (반영 대상)")
        mod = prompt_file(base_dir, "수정본 (제어파트, 기준)")
        hl = prompt_yes_no("변경 셀을 빨강으로 표시할까요?", default=True)
        op_merge(base_dir, orig, mod, hl)
    else:
        print("  ✗ 잘못된 선택입니다. 0~3 중에서 고르세요.")


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stdin.reconfigure(encoding="utf-8")
    except Exception:
        pass

    base_dir = get_base_dir()
    print(f"작업 폴더: {base_dir}")

    while True:
        print(BANNER)
        choice = input("선택 > ").strip()
        if choice in ("0", "q", "quit", "exit", "종료"):
            break
        try:
            run_menu_once(base_dir, choice)
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"\n✗ 처리 중 오류가 발생했습니다: {e}")
        print("\n" + "─" * 50)

    print("종료합니다.")


if __name__ == "__main__":
    try:
        main()
    except Exception:
        import traceback
        traceback.print_exc()
    # exe 더블클릭 시 창이 바로 닫히지 않도록 대기
    try:
        input("\nEnter 키를 누르면 창이 닫힙니다...")
    except EOFError:
        pass
