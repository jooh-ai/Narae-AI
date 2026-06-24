"""
BaroForecast - Seoul Air Base
HTML에서 토큰 자동 추출 후 API 호출 (브라우저/드라이버 불필요)
"""

import urllib.request          # 인터넷에서 데이터를 받아오는 표준 라이브러리
import json                    # JSON 문자열 <-> 파이썬 객체 변환
import re                      # 정규표현식: 텍스트에서 특정 패턴 찾기
from datetime import datetime, timedelta, timezone  # 날짜/시간 처리
import statistics              # 중앙값(median) 등 통계 계산
import openpyxl                # 엑셀(.xlsx) 파일 생성
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # 셀 꾸미기
from openpyxl.utils import get_column_letter  # 열 번호(1,2,..)를 문자(A,B,..)로 변환
import sys                     # 오류 발생 시 프로그램 강제 종료용

# ===== 전역 설정값(상수) =====
SPOT_ID = "kr134"                       # 윈드파인더 내부의 '서울공항' 지역 고유 ID
KST = timezone(timedelta(hours=9))      # 한국 표준시(UTC+9) 시간대
TARGET_HOURS = [0, 3, 6, 9, 12, 15, 18, 21]  # 추출할 시각(3시간 간격 8개)
MAX_DAYS = 7                            # 최대 7일치 예보만 사용
OUTPUT_FILE = f"{datetime.now().strftime('%Y-%m-%d')}.xlsx"  # 저장 파일명(오늘 날짜)
FORECAST_PAGE = "https://www.windfinder.com/forecast/seongnam_seoul_airbase"  # 예보 페이지 주소

# 일반 크롬 브라우저로 위장하기 위한 HTTP 요청 헤더(봇 차단 회피용)
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def fetch_token():
    """HTML에서 meta[name=version] 토큰 자동 추출"""
    # 예보 페이지에 접속하여 HTML 소스 전체를 텍스트로 받아온다
    req = urllib.request.Request(FORECAST_PAGE, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=15) as resp:
        html = resp.read().decode("utf-8")

    # <meta name="version" content="토큰"> 형태에서 토큰 값을 찾는다
    m = re.search(r'<meta\s+name=["\']version["\']\s+content=["\']([^"\']+)["\']', html)
    if not m:
        # 속성 순서가 반대인 경우(content가 먼저)도 대비
        m = re.search(r'<meta\s+content=["\']([^"\']+)["\']\s+name=["\']version["\']', html)
    if not m:
        # 두 경우 모두 실패하면 오류 출력 후 종료
        print("[오류] 토큰을 찾을 수 없습니다.")
        sys.exit(1)

    token = m.group(1)  # 정규식 괄호로 잡은 실제 토큰 문자열
    print(f"토큰 추출 성공: {token}")
    return token


def fetch_forecasts(token):
    # 예보 데이터를 제공하는 비공개 API 주소(최대 56개 항목 요청)
    url = f"https://api.windfinder.com/v2/spots/{SPOT_ID}/forecasts?limit=56"
    headers = {
        "User-Agent": HEADERS["User-Agent"],
        "Accept": "application/json",
        "Origin": "https://www.windfinder.com",
        "Referer": "https://www.windfinder.com/",
        # 추출한 토큰을 인증 헤더에 담아 접근 권한을 증명
        "Wf-Api-Authorization": f"WF-AUTH wfweb:1.0:{token}",
    }
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=15) as resp:
        # 응답(JSON 텍스트)을 파이썬 객체로 변환하여 반환
        return json.loads(resp.read().decode("utf-8"))


def fetch_update_time(token):
    # 예보가 마지막으로 갱신된 시각을 가져오는 보조 기능
    try:
        url = f"https://api.windfinder.com/v3/currentconditions/{SPOT_ID}"
        headers = {
            "User-Agent": HEADERS["User-Agent"],
            "Accept": "application/json",
            "Origin": "https://www.windfinder.com",
            "Referer": "https://www.windfinder.com/",
            "Wf-Api-Authorization": f"WF-AUTH wfweb:1.0:{token}",
        }
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        # cc -> 지역ID -> dtl(갱신 시각) 값을 안전하게 꺼낸다
        dtl = data.get("cc", {}).get(SPOT_ID, {}).get("dtl", "")
        if dtl:
            # ISO 시간 문자열을 한국 시간 "HH:MM" 형식으로 변환
            return datetime.fromisoformat(dtl).astimezone(KST).strftime("%H:%M")
    except Exception:
        # 갱신 시각은 부가 정보이므로 실패해도 무시
        pass
    return "N/A"


def parse_forecasts(items):
    result = {}  # {날짜: {시각: {기온, 기압}}} 형태로 정리할 딕셔너리
    for item in items:
        dtl = item.get("dtl")  # 각 예보 항목의 시간 정보
        if not dtl:
            continue  # 시간 정보 없으면 건너뜀
        try:
            dt_kst = datetime.fromisoformat(dtl).astimezone(KST)  # 한국 시간으로 변환
        except Exception:
            continue  # 변환 실패 시 건너뜀

        hour = dt_kst.hour
        if hour not in TARGET_HOURS:
            continue  # 원하는 3시간 간격 시각이 아니면 건너뜀

        # 영어 요일을 한국어로 변환하기 위한 매핑
        DAY_KO = {"Monday":"월요일","Tuesday":"화요일","Wednesday":"수요일","Thursday":"목요일","Friday":"금요일","Saturday":"토요일","Sunday":"일요일"}
        # "월요일, 6월 4일" 형태의 날짜 문자열 생성
        date_str = f"{DAY_KO[dt_kst.strftime('%A')]}, {dt_kst.month}월 {dt_kst.day}일"
        pressure = round(float(item["ap"])) if item.get("ap") is not None else None  # ap=대기압(반올림)
        temp     = round(float(item["at"])) if item.get("at") is not None else None  # at=기온(반올림)

        # 날짜별, 시각별로 값을 저장
        if date_str not in result:
            result[date_str] = {}
        result[date_str][hour] = {"temperature": temp, "pressure": pressure}

    def parse_date(d):
        # 날짜 정렬을 위해 "6월 4일"에서 (월, 일) 숫자 쌍을 추출
        try:
            # "월요일, 3월 17일" 형식 파싱
            m = re.search(r'(\d+)월 (\d+)일', d)
            if m:
                return (int(m.group(1)), int(m.group(2)))
        except Exception:
            pass
        return (99, 99)  # 파싱 실패 시 맨 뒤로 보내는 기본값

    # (월, 일) 기준으로 정렬한 뒤 최대 7일치만 추려서 반환
    sorted_dates = sorted(result.keys(), key=parse_date)[:MAX_DAYS]
    return {d: result[d] for d in sorted_dates}


def make_excel(data, update_time, output_path):
    wb = openpyxl.Workbook()  # 새 엑셀 통합문서 생성
    ws = wb.active            # 기본 시트 선택
    ws.title = datetime.now().strftime("%Y-%m-%d")  # 시트 이름을 오늘 날짜로

    # 미리 정의해두는 셀 스타일들
    YELLOW_FILL = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")  # 노란 배경
    CENTER = Alignment(horizontal="center", vertical="center")  # 가운데 정렬
    thin = Side(style="thin")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)  # 얇은 사방 테두리

    def sc(cell, bold=False, fill=None):
        # 셀에 가운데정렬+테두리를 기본 적용하고, 옵션으로 굵게/배경색 지정
        cell.alignment = CENTER
        cell.border = BORDER
        if bold:
            cell.font = Font(bold=True)
        if fill:
            cell.fill = fill

    # 상단 정보 영역 작성
    ws["A1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # 파일 생성 시각
    ws["A2"] = "Update time :"
    ws["B2"] = update_time                                   # 예보 갱신 시각
    ws["A3"] = "Seoul Air Base"                              # 지역 이름

    sorted_dates = list(data.keys())  # 표에 그릴 날짜 목록(정렬된 상태)

    def write_section(start_row, title, key, last_col_label):
        # 표 제목을 노란 배경 + 굵은 글씨로 작성
        c = ws.cell(row=start_row, column=1, value=title)
        c.fill = YELLOW_FILL
        c.font = Font(bold=True)
        c.alignment = CENTER

        # 헤더 행 작성: Local time | 시각들 | 중위 | 최소/취약
        hr = start_row + 1
        sc(ws.cell(row=hr, column=1, value="Local time"), bold=True)
        for ci, h in enumerate(TARGET_HOURS, start=2):  # B열(2)부터 시각 나열
            sc(ws.cell(row=hr, column=ci, value=h), bold=True)
        sc(ws.cell(row=hr, column=10, value="중위"), bold=True)            # J열: 중앙값
        sc(ws.cell(row=hr, column=11, value=last_col_label), bold=True)   # K열: 최소/취약

        # 날짜별로 한 행씩 데이터 채우기
        for ri, date_str in enumerate(sorted_dates):
            row = hr + 1 + ri
            sc(ws.cell(row=row, column=1, value=date_str))  # A열: 날짜
            row_vals = []        # 그 날 전체 값 목록
            afternoon_vals = []  # 12시 이후(오후) 값 목록
            for ci, hour in enumerate(TARGET_HOURS, start=2):
                val = data.get(date_str, {}).get(hour, {}).get(key)  # 해당 시각의 값
                sc(ws.cell(row=row, column=ci, value=val))
                if val is not None:
                    row_vals.append(val)
                    if hour >= 12:
                        afternoon_vals.append(val)

            # 중앙값(중위) 계산 후 J열에 굵게 표시
            med = round(statistics.median(row_vals)) if row_vals else ""
            c_med = ws.cell(row=row, column=10, value=med)
            c_med.alignment = CENTER; c_med.border = BORDER; c_med.font = Font(bold=True)

            # 마지막 열 값: 기온표("취약")는 오후 최솟값, 그 외는 하루 전체 최솟값
            ext_vals = afternoon_vals if (last_col_label == "취약" and afternoon_vals) else row_vals
            ext = round(min(ext_vals)) if ext_vals else ""
            c_ext = ws.cell(row=row, column=11, value=ext)
            c_ext.alignment = CENTER; c_ext.border = BORDER; c_ext.font = Font(bold=True)

        # 다음 표가 시작될 행 번호를 반환
        return hr + 1 + len(sorted_dates)

    # 기압표 작성 후, 그 아래에 기온표 작성
    next_row = write_section(5, "Pressure", "pressure", "최소")
    write_section(next_row + 1, "Tempereture", "temperature", "취약")

    # 열 너비 조정: 날짜 열(A)은 넓게, 숫자 열(B~K)은 좁게
    ws.column_dimensions["A"].width = 22
    for i in range(2, 12):
        ws.column_dimensions[get_column_letter(i)].width = 8

    wb.save(output_path)  # 엑셀 파일 저장
    print(f"[완료] 저장됨: {output_path}")


def main():
    # 전체 작업을 순서대로 실행하는 진입점
    print("윈드파인더에서 정보를 취득하고 있습니다.")

    token = fetch_token()                    # ① HTML에서 토큰 추출
    update_time = fetch_update_time(token)   # ② 예보 갱신 시각 조회
    print(f"업데이트 시간: {update_time}")

    items = fetch_forecasts(token)           # ③ 예보 데이터 받아오기
    print(f"수신된 항목 수: {len(items)}")

    data = parse_forecasts(items)            # ④ 데이터 가공/정렬
    print(f"파싱된 날짜 수: {len(data)}일")

    make_excel(data, update_time, OUTPUT_FILE)  # ⑤ 엑셀 파일 생성


# 이 파일을 직접 실행할 때만 main() 호출
if __name__ == "__main__":
    main()
