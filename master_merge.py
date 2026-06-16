# -*- coding: utf-8 -*-
"""
자재마스터 병합 도구 (수정본 → 원본 반영)
================================================

수정 파일(제어파트만 추출하여 작업한 파일)의 데이터를, A열(자재) 키로
원본 파일의 같은 자재 행에 반영합니다.

- 드라이버는 '수정 파일'입니다. 수정 파일에 있는 각 행(=제어파트)을
  A열 키로 원본에서 찾아 갱신합니다. 수정 파일에 없는 자재(타 파트)는
  원본에서 건드리지 않습니다.
  → 원본에서 다른 파트였다가 수정 파일에서 제어파트로 바뀐 행도,
    수정 파일에 들어있으므로 자연히 갱신됩니다(H열도 갱신).

반영 규칙
---------
- B, H, J, K, L, N : 수정 파일에 '값이 있을 때만' 덮어씀.
                    (수정이 빈 셀이면 원본 값을 그대로 유지 — 휴먼에러 방지)
- O ~ AH (개별속성1~20) : '빈 셀도' 덮어씀(완전 미러).
                          수정이 비어있으면 원본도 비움.

- 셀 서식(폰트·색상·테두리·열 너비 등)은 그대로 보존합니다.
- 실제로 값이 '변경된' 셀만 빨간 글자색으로 표시합니다.
- 원본을 건드리지 않고 결과를 새 파일로 저장합니다(동일 폴더).

사용법
------
1) 한 번만:  터미널에서  ->  pip install openpyxl
2) 아래 ##### 설정 ##### 블록의 파일명을 정확히 맞춤
3) VSCode에서 F5(또는 '실행')로 이 파일을 실행
"""

import os
import sys
import re
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import column_index_from_string
except ImportError:
    print("✗ openpyxl 라이브러리가 필요합니다. 터미널에서 다음을 실행하세요:")
    print("    pip install openpyxl")
    sys.exit(1)


# ═══════════════════════════════════════════════
#                    설정
# ═══════════════════════════════════════════════
# 원본 파일 (전체 파트 포함, 여기에 반영하여 새 파일로 저장)
ORIGINAL_FILE = "사업소 자재마스터 _26.04.29_위례.xlsx"

# 수정 파일 (제어파트만 추출·작업한 파일, 반영의 기준)
MODIFIED_FILE = "사업소 자재마스터 _26.06.15_위례_제어.xlsx"

# 두 파일의 시트명 (동일)
SHEET_NAME = "4.위례"

# 헤더 행 번호 (이 행 다음부터 데이터)
HEADER_ROW = 1

# 자재(키) 컬럼
KEY_COL = "A"

# 값이 있을 때만 덮어쓰는 열 (수정이 빈 셀이면 원본 유지)
VALUE_COLS = ["B", "H", "J", "K", "L", "N"]

# 빈 셀도 덮어쓰는 열(완전 미러) 범위: O ~ AH (개별속성1~20)
MIRROR_COL_START = "O"
MIRROR_COL_END = "AH"

# 변경된 셀을 빨간 글자색으로 강조할지 여부
HIGHLIGHT_RED = True

# 결과 파일명. 빈 문자열("")이면 "원본파일명_merged_YYMMDD_HHMM.xlsx" 로 자동 생성
OUTPUT_FILE = ""
# ═══════════════════════════════════════════════


RED_ARGB = "FFFF0000"


# ─────────────────────────────────────────────
# 헬퍼
# ─────────────────────────────────────────────
def cell_text(v):
    """셀 '값'을 문자열로 (None → '')."""
    if v is None:
        return ""
    return str(v)


def is_empty(v):
    if v is None:
        return True
    s = str(v)
    return re.sub(r"[\s ​‌‍﻿]+", "", s) == ""


def key_norm(v):
    """자재(키) 비교용 정규화: 앞뒤 공백 제거, 정수형 실수는 정수 문자열로."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def same_value(a, b):
    """두 셀 값이 사실상 같은지 (사용자 시각)."""
    return cell_text(a).strip() == cell_text(b).strip()


def apply_red_font(cell):
    f = cell.font
    cell.font = Font(
        name=f.name, size=f.size, bold=f.bold, italic=f.italic,
        vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
        color=RED_ARGB,
    )


def resolve_path(base_dir, path):
    return path if os.path.isabs(path) else os.path.join(base_dir, path)


def get_sheet(wb, name, label):
    if name not in wb.sheetnames:
        print(f"✗ {label}에서 시트 '{name}'를 찾을 수 없습니다. 사용 가능: {wb.sheetnames}")
        sys.exit(1)
    return wb[name]


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def main():
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        base_dir = os.getcwd()

    orig_path = resolve_path(base_dir, ORIGINAL_FILE)
    mod_path = resolve_path(base_dir, MODIFIED_FILE)
    for label, p in [("원본", orig_path), ("수정", mod_path)]:
        if not os.path.exists(p):
            print(f"✗ {label} 파일을 찾을 수 없습니다: {p}")
            print("  설정 블록의 파일 경로를 확인하세요.")
            sys.exit(1)

    key_c = column_index_from_string(KEY_COL)
    value_cols = [column_index_from_string(c) for c in VALUE_COLS]
    mir_start = column_index_from_string(MIRROR_COL_START)
    mir_end = column_index_from_string(MIRROR_COL_END)
    if mir_end < mir_start:
        print("✗ 미러 열 범위(O~AH) 설정을 확인하세요.")
        sys.exit(1)

    # 원본: 편집·저장 대상이므로 일반 로드(서식/수식 보존)
    print(f"▶ 원본 로드 중: {orig_path}")
    keep_vba = orig_path.lower().endswith(".xlsm")
    owb = load_workbook(orig_path, keep_vba=keep_vba)
    ows = get_sheet(owb, SHEET_NAME, "원본")

    # 수정: 값만 읽으면 되므로 data_only 로드
    print(f"▶ 수정 로드 중: {mod_path}")
    mwb = load_workbook(mod_path, data_only=True)
    mws = get_sheet(mwb, SHEET_NAME, "수정")

    # ── 원본 A열 → 행번호 인덱스 ──
    orig_index = {}
    dup_orig = 0
    o_max = ows.max_row
    for r in range(HEADER_ROW + 1, o_max + 1):
        k = key_norm(ows.cell(row=r, column=key_c).value)
        if k == "":
            continue
        if k in orig_index:
            dup_orig += 1
            continue   # 첫 행 유지 (자재 코드는 보통 유일)
        orig_index[k] = r
    print(f"  · 원본 자재 {len(orig_index)}종 인덱싱" +
          (f" (중복 자재 {dup_orig}건은 첫 행만 사용)" if dup_orig else ""))

    # ── 수정 파일을 돌며 원본에 반영 ──
    m_max = mws.max_row
    matched = 0          # 원본에서 찾아 갱신한 행
    not_found = 0        # 원본에 없는 자재(수정에만 존재)
    empty_key = 0        # 수정 A열이 비어있는 행
    value_changed = 0    # B/H/J/K/L/N 변경 셀
    mirror_changed = 0   # O~AH 변경 셀(기입)
    mirror_cleared = 0   # O~AH 변경 셀(비움)
    h_changed = 0        # H열(관리부서)이 바뀐 행
    sample_not_found = []
    sample_h = []
    h_col = column_index_from_string("H")

    for mr in range(HEADER_ROW + 1, m_max + 1):
        k = key_norm(mws.cell(row=mr, column=key_c).value)
        if k == "":
            empty_key += 1
            continue
        orow = orig_index.get(k)
        if orow is None:
            not_found += 1
            if len(sample_not_found) < 15:
                sample_not_found.append(k)
            continue
        matched += 1

        # 값이 있을 때만 덮어쓰는 열 (B/H/J/K/L/N)
        for col in value_cols:
            mv = mws.cell(row=mr, column=col).value
            if is_empty(mv):
                continue
            ocell = ows.cell(row=orow, column=col)
            if not same_value(ocell.value, mv):
                old = ocell.value
                ocell.value = mv
                if HIGHLIGHT_RED:
                    apply_red_font(ocell)
                value_changed += 1
                if col == h_col:
                    h_changed += 1
                    if len(sample_h) < 15:
                        sample_h.append(f"{k}: '{cell_text(old)}'→'{cell_text(mv)}' (원본 {orow}행)")

        # 빈 셀도 덮어쓰는 열 (O~AH 완전 미러)
        for col in range(mir_start, mir_end + 1):
            mv = mws.cell(row=mr, column=col).value
            ocell = ows.cell(row=orow, column=col)
            if same_value(ocell.value, mv):
                continue
            if is_empty(mv):
                ocell.value = None        # 원본을 비움
                mirror_cleared += 1
            else:
                ocell.value = mv
                if HIGHLIGHT_RED:
                    apply_red_font(ocell)
                mirror_changed += 1

        if (mr - HEADER_ROW) % 500 == 0:
            print(f"  ... 수정 {mr - HEADER_ROW}/{m_max - HEADER_ROW} 행 처리")

    print(f"  → 병합 완료: 매칭·갱신 {matched}행, 원본에 없음 {not_found}행, A열 비어있음 {empty_key}행")
    print(f"     · B/H/J/K/L/N 변경 {value_changed}칸 (그중 H열 부서 변경 {h_changed}행)")
    print(f"     · O~AH 변경 {mirror_changed}칸 기입 + {mirror_cleared}칸 비움")
    if h_changed:
        print("     · 부서(H열) 변경 내역:")
        for s in sample_h:
            print(f"         - {s}")
    if not_found:
        print(f"  ⚠ 수정 파일에만 있고 원본에 없는 자재 {not_found}건 (반영 못함):")
        for s in sample_not_found:
            print(f"         - {s}")

    # ── 저장 ──
    if OUTPUT_FILE:
        out_path = resolve_path(base_dir, OUTPUT_FILE)
    else:
        base = os.path.splitext(os.path.basename(orig_path))[0]
        stamp = datetime.now().strftime("%y%m%d_%H%M")
        ext = ".xlsm" if keep_vba else ".xlsx"
        out_path = os.path.join(base_dir, f"{base}_merged_{stamp}{ext}")

    owb.save(out_path)
    print(f"✅ 저장 완료: {out_path}")


if __name__ == "__main__":
    main()
