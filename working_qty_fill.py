# -*- coding: utf-8 -*-
"""
Working Q'ty 채우기 도구 (N열)
================================================

자재마스터 파일의 A열(자재)과 재고 파일의 A열(자재)을 비교하여,
같은 자재끼리 매칭한 뒤 재고 파일의 R열(max) 값을 자재마스터의
N열(Working Q'ty)에 기입합니다.

- N열 기입 외의 다른 처리(개별속성/제조사/모델명/MAKER P/N/사양 등)는
  전혀 하지 않습니다.
- 셀 서식(폰트·색상·테두리·열 너비 등)은 그대로 보존됩니다.
- 결과는 원본을 건드리지 않고 새 파일로 저장합니다.

재고 파일에서 같은 A열(자재) 값이 여러 행에 중복되어 있으면,
그 행들의 R열(max) 값 중 '가장 큰 값'을 사용합니다.

사용법
------
1) 한 번만:  터미널에서  ->  pip install openpyxl
2) 아래 ##### 설정 ##### 블록의 값을 본인 파일에 맞게 수정
3) VSCode에서 F5(또는 '실행')로 이 파일을 실행
"""

import os
import sys
import re
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import column_index_from_string
except ImportError:
    print("✗ openpyxl 라이브러리가 필요합니다. 터미널에서 다음을 실행하세요:")
    print("    pip install openpyxl")
    sys.exit(1)


# ═══════════════════════════════════════════════
#                    설정
# ═══════════════════════════════════════════════
# 수정 대상: 자재마스터 파일 (N열을 채울 파일)
MASTER_FILE       = "사업소 자재마스터 _26.06.04_위례_제어_260605.xlsx"

# 비교용: 재고 파일
INVENTORY_FILE    = "재고.xlsx"

# 시트명. 빈 문자열("")이면 자재마스터는 '위례' 포함 시트(없으면 첫 시트),
#                          재고는 첫 시트를 자동 선택
MASTER_SHEET      = ""
INVENTORY_SHEET   = ""

# 헤더 행 번호 (이 행 다음부터 데이터로 처리)
MASTER_HEADER_ROW    = 1
INVENTORY_HEADER_ROW = 1

# 컬럼 지정 (엑셀 알파벳)
MASTER_KEY_COL    = "A"   # 자재마스터의 자재(키) 컬럼
MASTER_TARGET_COL = "N"   # 자재마스터의 Working Q'ty (기입 대상) 컬럼
INVENTORY_KEY_COL = "A"   # 재고의 자재(키) 컬럼
INVENTORY_VAL_COL = "R"   # 재고의 max (가져올 값) 컬럼

# 매칭된 행의 N열에 새로 기입한 값을 빨간 글자색으로 강조할지 여부
HIGHLIGHT_RED     = True

# 이미 N열에 값이 있어도 매칭되면 덮어쓸지 여부
#   True  : 매칭되면 항상 재고 값으로 덮어씀
#   False : N열이 비어있는 경우에만 기입 (기존 값 보존)
OVERWRITE_EXISTING = True

# 결과 파일명. 빈 문자열("")이면 "자재마스터파일명_WorkingQty_YYMMDD_HHMM.xlsx" 로 자동 생성
OUTPUT_FILE       = ""
# ═══════════════════════════════════════════════


RED_ARGB = "FFFF0000"


# ─────────────────────────────────────────────
# 헬퍼
# ─────────────────────────────────────────────
def cell_text(cell):
    """셀 값을 문자열로 반환 (None → '')."""
    v = cell.value
    if v is None:
        return ""
    return str(v)


def is_empty(v):
    """공백/제로폭 문자만 있으면 빈 값으로 간주."""
    if v is None:
        return True
    s = str(v)
    cleaned = re.sub(r"[\s ​‌‍﻿]+", "", s)
    return cleaned == ""


def key_norm(v):
    """자재(키) 비교용 정규화.
    - 앞뒤 공백 제거
    - 숫자로 저장된 정수형(예: 12345.0)은 '12345'로 통일
    - 그 외는 좌우 공백만 제거한 문자열"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def to_number(v):
    """R열(max) 값을 비교용 숫자로 변환. 변환 불가 시 None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def apply_red_font(cell):
    """기존 폰트 속성을 보존하면서 글자색만 빨강으로 변경."""
    f = cell.font
    cell.font = Font(
        name=f.name, size=f.size, bold=f.bold, italic=f.italic,
        vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
        color=RED_ARGB,
    )


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
def resolve_path(base_dir, path):
    return path if os.path.isabs(path) else os.path.join(base_dir, path)


def pick_sheet(wb, preferred):
    if preferred:
        if preferred not in wb.sheetnames:
            print(f"✗ 시트 '{preferred}'를 찾을 수 없습니다. 사용 가능한 시트: {wb.sheetnames}")
            sys.exit(1)
        return wb[preferred]
    chosen = next((n for n in wb.sheetnames if "위례" in n), wb.sheetnames[0])
    return wb[chosen]


def main():
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        base_dir = os.getcwd()

    master_path = resolve_path(base_dir, MASTER_FILE)
    inv_path = resolve_path(base_dir, INVENTORY_FILE)

    for label, p in [("자재마스터", master_path), ("재고", inv_path)]:
        if not os.path.exists(p):
            print(f"✗ {label} 파일을 찾을 수 없습니다: {p}")
            print("  설정 블록의 파일 경로를 확인하세요.")
            sys.exit(1)

    mkey = column_index_from_string(MASTER_KEY_COL)
    mtgt = column_index_from_string(MASTER_TARGET_COL)
    ikey = column_index_from_string(INVENTORY_KEY_COL)
    ival = column_index_from_string(INVENTORY_VAL_COL)

    # ── 재고 파일에서 자재(A) → max(R) 맵 구성 (중복 시 가장 큰 값) ──
    print(f"▶ 재고 파일 로드 중: {inv_path}")
    inv_wb = load_workbook(inv_path, data_only=True)
    inv_ws = pick_sheet(inv_wb, INVENTORY_SHEET)
    print(f"  · 재고 시트: '{inv_ws.title}'")

    inv_map = {}          # key_norm → (number, original_value)
    dup_keys = set()      # 중복이 발견된 키
    inv_rows = 0
    inv_skipped = 0       # R값이 숫자가 아니어서 건너뛴 행
    inv_max_row = inv_ws.max_row
    for r in range(INVENTORY_HEADER_ROW + 1, inv_max_row + 1):
        k = key_norm(inv_ws.cell(row=r, column=ikey).value)
        if k == "":
            continue
        raw = inv_ws.cell(row=r, column=ival).value
        num = to_number(raw)
        if num is None:
            inv_skipped += 1
            continue
        inv_rows += 1
        if k in inv_map:
            dup_keys.add(k)
            if num > inv_map[k][0]:
                inv_map[k] = (num, raw)   # 더 큰 R값으로 갱신
        else:
            inv_map[k] = (num, raw)
    print(f"  · 재고 자재 {len(inv_map)}종 인덱싱 (유효 행 {inv_rows}건, "
          f"중복 자재 {len(dup_keys)}종, R값 숫자아님 {inv_skipped}건 제외)")

    # ── 자재마스터 로드 후 N열 채우기 ──
    print(f"▶ 자재마스터 로드 중: {master_path}")
    keep_vba = master_path.lower().endswith(".xlsm")
    wb = load_workbook(master_path, keep_vba=keep_vba)
    ws = pick_sheet(wb, MASTER_SHEET)
    print(f"  · 자재마스터 시트: '{ws.title}'")

    max_row = ws.max_row
    written = 0       # N열에 기입한 행
    no_match = 0      # 재고에 없는 자재
    skipped_existing = 0  # N열에 값이 있어 건너뜀 (OVERWRITE_EXISTING=False)
    empty_key = 0     # A열이 비어있는 행

    for r in range(MASTER_HEADER_ROW + 1, max_row + 1):
        k = key_norm(ws.cell(row=r, column=mkey).value)
        if k == "":
            empty_key += 1
            continue
        hit = inv_map.get(k)
        if hit is None:
            no_match += 1
            continue
        tgt_cell = ws.cell(row=r, column=mtgt)
        if not OVERWRITE_EXISTING and not is_empty(tgt_cell.value):
            skipped_existing += 1
            continue
        tgt_cell.value = hit[1]   # 재고 R열(max) 원본 값
        if HIGHLIGHT_RED:
            apply_red_font(tgt_cell)
        written += 1

        if (r - MASTER_HEADER_ROW) % 1000 == 0:
            print(f"  ... {r - MASTER_HEADER_ROW}/{max_row - MASTER_HEADER_ROW} 행 처리")

    print(f"  → 완료: N열 기입 {written}행, 재고에 없음 {no_match}행, "
          f"A열 비어있음 {empty_key}행" +
          ("" if OVERWRITE_EXISTING else f", 기존값 보존 {skipped_existing}행"))

    # ── 저장 ──
    if OUTPUT_FILE:
        out_path = resolve_path(base_dir, OUTPUT_FILE)
    else:
        base = os.path.splitext(os.path.basename(master_path))[0]
        stamp = datetime.now().strftime("%y%m%d_%H%M")
        ext = ".xlsm" if keep_vba else ".xlsx"
        out_path = os.path.join(base_dir, f"{base}_WorkingQty_{stamp}{ext}")

    wb.save(out_path)
    print(f"✅ 저장 완료: {out_path}")


if __name__ == "__main__":
    main()
